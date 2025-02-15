#开发者：
#要实现的功能:
## 0.以下所有东西，控件都用TKinter实现
## 1.实现单本图书的手动借,还(用学生码)
## 2.实现单本图书的库存添加(+？)
## 3.实现单本图书的删除(查询,DELETE)
## 4.实现excel导入图书信息()-|
###|->4.0.弹出新窗口
###|->4.1.添加excel文件选择按钮
###|->4.2.预览表格(1~2条)(4.3, 4.4每个修改时同步更新)
###|->4.3.1.书籍信息指定列(书名，作者，出版社，isbn，书籍介绍，库存)
###|->4.3.2.单选按钮 "是否每行书名，作者，出版社，isbn，书籍介绍，库存中有一个为空时跳过?",默认选择
###|->4.4.书籍信息从第几行开始添加(默认1)
###|->4.5.“导入”按钮
###|->4.6.“提示导入中，请等待”，等待导入程序运行完就提示成功并关闭导入窗口
## 5.实现清空图书信息
## 6.实现单人信息添加->
###|->6.1.姓名
###|->6.2.学号
###|->6.3.班级
###|-|->6.4.添加按钮
## 7.实现单人信息注销->
###|->7.1.姓名
###|->7.2.学号
###|->7.3.班级
## 8.借阅信息查询按钮，点击打开新窗口“查询分类”，这个窗口的有3个按钮，一个是"正在借阅的学生"，一个是"图书馆借过的书"，一个是“返回”，前2个按钮点击后会弹出不同的2个窗口，返回按钮点击会返回到第一个页面-|
###|->8.1.0 窗口1“正在借阅的学生”
###|->8.1.1 窗口1中第一行有一个实时的时间(年月日时分秒，居中显示)，第二行有一个返回按钮，一个按钮“全部”，一个按钮“已逾期”，一个按钮“导出表格内容”，这些按钮都在窗口上面第二行，下面部分是一个大表格，充满窗口
###|->8.1.2 表格中有4列，分别是“学号”，“姓名”，“班级”，“借书时间”，“应还时间”，“书名”，“作者”，“出版社”，“isbn”
####|->8.1.2.1 当鼠标点击某一行时，弹出窗口，上面询问是否删除信息，点击“确认”后删除该行信息以及数据库信息，点击“取消”则不删除
###|->8.1.2.3 窗口1“正在借阅的学生”默认显示全部信息，当点击“已逾期”时，显示“应还时间”早于当前时间的学生信息并显示在窗口中，点击“全部”时，显示全部学生信息，不用管逾期不逾期
####|->8.1.2.3 当鼠标点击某一行时，弹出窗口，上面询问是否删除信息，点击“确认”后删除该行信息以及数据库信息，点击“取消”则不删除
###|->8.1.3 当点击“导出表格内容”时，让用户选择导出路径，然后导出表格内容到excel文件中(含表头，即“学号”，“姓名”，“班级”，“借书时间”，“应还时间”，“书名”，“作者”，“出版社”，“isbn”)，文件名为“正在借阅的学生.xlsx”,保存到指定路径中
###|->8.2.0 窗口2“图书馆借出过的书”
###|->8.2.1 窗口中有一个返回按钮，一个按钮“导出表格内容”，第二行有一个输入框用来搜索的，下面是一个大表格，充满窗口
###|->8.2.2 表格中有5列，分别是“书名”，“作者”，“出版社”，“出版时间”，“isbn”
###|->8.2.3 实时监测输入框的新输入信息，查询表格中“书名”1列中是否有符合的书籍信息，并计算数量显示到提示框，全部显示在表格中，如果输入内容为空，则显示全部信息
###|->8.2.4 当点击“导出表格内容”时，让用户选择导出路径，然后导出表格内容到excel文件中(含表头，即“书名”，“作者”，“出版社”，“出版时间”，“isbn”)，文件名为“图书馆借出过的书.xlsx”,保存到指定路径中
## 9.一个按钮“导出日志”->
###|->9.1点击按钮后创建新的窗口，一个返回按钮，一个导出日志按钮，下面有一个表格，充满窗口
###|->9.2下面的表格里有2列，“操作时间”，“操作”
###|->9.3当点击“导出表格内容”时，让用户选择导出路径，然后导出表格内容到excel文件中(含表头，即“操作时间”，“操作”)，文件名为“操作日志.xlsx”,保存到指定路径中
##这样连接数据库Librarysql = sqlite3.connect(mypath("Library.oflibrary"))systemlog = sqlite3.connect(mypath("Library.log"))
###数据库Library.oflibrary的内容:
####Create table books(bookname text,author text,press text,publicationTime text,bookInfo text,isbn text,inventory int);
####Create table borrow(bookname text,author text,press text,publicationTime text,isbn text,borrowtime text,MustReturnTime text,ReaderName text,Readerid int,Readerclass int);
####Create table borrowhistory(bookname text,author text,press text,publicationTime text,isbn text,borrowtime text,MustBookReturnTime text,TrueBookReturnTime text,ReaderName text,Readerid int,Readerclass int);
####Create table users(Username text,Userid int,Userclass int,UserBorrowBooks text,UserPassword text,UserBorrowedBooks text);
###数据库Library.log的内容:
####Create table log(Time text,Do text);
##添加这个代码在开头：
###def mypath(other: str | None = ""):
# return str(os.path.dirname(os.path.abspath(__file__)))+"\\"+other
import os
import sqlite3
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import pandas as pd
from datetime import datetime
import maliang
import openpyxl
from library import lb
from library import *
import time
from tkinter import font
import re

# global Librarysql
# global systemlog
# global cursor1
# global cursor2

def mypath(other: str | None = ""):
    return os.path.dirname(os.path.abspath(__file__)) + "\\" + other

class LibrarySystem(maliang.Tk):
    def __init__(self):
        self.excel_import_start_row = "2"
        self.excel_import_continue_None = True
        self.excel_import_bookname_column = "2"
        self.excel_import_author_column = "4"
        self.excel_import_press_column = "6"
        self.excel_import_publicationTime_column = "7"
        self.excel_import_bookInfo_column = "11"
        self.excel_import_isbn_column = "5"
        self.excel_import_inventory_column = "13"
        self.excel_max_column = 7
        self.excel_max_row = 2
        self.file_path = None
        self.one_borrow_search_type = None
        self.now_search_borrow_books_list = []
        self.want_to_borrow_books_list = []
        self.want_to_borrow_books_list_name_and_isbn_tuple = []
        self.borrow_student_name = ""
        self.borrow_student_id = ""
        self.borrow_student_class = ""
        self.borrow_student_password = ""
        self.borrow_student_borrow_books = ""
        self.borrow_student_borrowed_books = ""
        self.one_return_search_type = None
        self.now_search_return_books_list = []
        self.want_to_return_books_list = []
        self.want_to_return_books_list_name_and_isbn_tuple = []
        self.return_student_name = ""
        self.return_student_id = ""
        self.return_student_class = ""
        self.return_student_password = ""
        self.return_student_borrow_books = ""
        self.return_student_borrowed_books = ""
        self.one_delete_book_search_type = None
        self.one_delete_book_now_show_books_list = []
        self._one_delete_book_book_name = ""
        self._one_delete_book_author_name = ""
        self._one_delete_book_press = ""
        self._one_delete_book_publicationTime = ""
        self._one_delete_book_bookInfo = ""
        self._one_delete_book_isbn = ""
        self._one_delete_book_inventory = ""
        self._one_delete_book_id = ""
        self.now_show_amend_books_list = []
        self.amend_book_bookname_first_msg = ""
        self.amend_book_author_first_msg = ""
        self.amend_book_press_first_msg = ""
        self.amend_book_publicationTime_first_msg = ""
        self.amend_book_bookInfo_first_msg = ""
        self.amend_book_isbn_first_msg = ""
        self.amend_book_inventory_first_msg = ""
        self.amend_book_id_first_msg = ""
        self.amend_book_bookname_second_msg = ""
        self.amend_book_author_second_msg = ""
        self.amend_book_press_second_msg = ""
        self.amend_book_publicationTime_second_msg = ""
        self.amend_book_bookInfo_second_msg = ""
        self.amend_book_isbn_second_msg = ""
        self.amend_book_inventory_second_msg = ""
        self.amend_book_id_second_msg = ""
        self.add_student_qrcode_save_path = "f:\\py\\mylibrarysystem\\学生信息\\"
        self.add_student_show_qrcode_when_finish = True
        self.delete_student_name = ""
        self.delete_student_class = ""
        self.delete_student_id = ""
        self.delete_student_password = ""
        self.delete_student_borrow_books = ""
        self.delete_student_borrowed_books = ""
        self.delete_student_now_show_students_list = []














        self.root = maliang.Tk(size=(800,600),title="图书馆终端")
        self.root.iconbitmap(mypath("favicon.ico"))
        self.root.center()
        self.head_menus = tk.Menu(self.root)

        self.system_menu = tk.Menu(self.head_menus, tearoff=0)
        self.system_menu.add_command(label="退出", command=lambda:exit())
        self.head_menus.add_cascade(label="系统", menu=self.system_menu)



        self.borrow_return_menu = tk.Menu(self.head_menus, tearoff=0)
        self.borrow_return_menu.add_command(label="单本书借阅", command=self.borrow_book)
        self.borrow_return_menu.add_command(label="单本书还书", command=self.return_book)
        self.borrow_return_menu.add_separator()
        self.borrow_return_menu.add_command(label="借阅信息总查询", command=self.borrow_and_return_query)
        self.head_menus.add_cascade(label="关于借阅", menu=self.borrow_return_menu)

        self.about_book_menu = tk.Menu(self.head_menus, tearoff=0)
        self.about_book_menu.add_command(label="Excel导入所有书籍", command=self.import_book)
        self.about_book_menu.add_command(label="清空所有书籍", command=self.delete_all_book)
        self.about_book_menu.add_separator()
        self.about_book_menu.add_command(label="删除单本图书", command=self.delete_one_book)
        self.about_book_menu.add_separator()
        self.about_book_menu.add_command(label="修改书籍信息", command=self.amend_book_info)
        self.head_menus.add_cascade(label="关于书籍", menu=self.about_book_menu)

        self.about_student_menu = tk.Menu(self.head_menus, tearoff=0)
        self.about_student_menu.add_command(label="添加学生信息", command=self.add_student_info)
        self.about_student_menu.add_command(label="注销学生信息", command=self.delete_student_info)
        self.about_student_menu.add_command(label="修改学生信息", command=self.amend_student_info)
        self.head_menus.add_cascade(label="关于学生", menu=self.about_student_menu)

        self.about_log_menu = tk.Menu(self.head_menus, tearoff=0)
        self.about_log_menu.add_command(label="导出日志", command=self.export_log)
        self.head_menus.add_cascade(label="其它", menu=self.about_log_menu)


        cv = maliang.Canvas(self.root,auto_zoom=True)
        cv.place(width=1280, height=720,x=0,y=0)
        self.now_time = maliang.Text(cv,(310,0),text=str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))


        self.now_splite_has_books_show_tree = ttk.Treeview(cv,columns=("书名","作者","出版社","出版时间","书籍介绍","ISBN","库存","id"),show="headings")
        self.now_splite_has_books_show_tree.heading("书名", text="书名")
        self.now_splite_has_books_show_tree.heading("作者", text="作者")
        self.now_splite_has_books_show_tree.heading("出版社", text="出版社")
        self.now_splite_has_books_show_tree.heading("出版时间", text="出版时间")
        self.now_splite_has_books_show_tree.heading("书籍介绍", text="书籍介绍")
        self.now_splite_has_books_show_tree.heading("ISBN", text="ISBN")
        self.now_splite_has_books_show_tree.heading("库存", text="库存")
        self.now_splite_has_books_show_tree.heading("id", text="id")
        self.now_splite_has_books_show_tree.column("书名", width=90)
        self.now_splite_has_books_show_tree.column("作者", width=90)
        self.now_splite_has_books_show_tree.column("出版社", width=80)
        self.now_splite_has_books_show_tree.column("出版时间", width=60)
        self.now_splite_has_books_show_tree.column("书籍介绍", width=80)
        self.now_splite_has_books_show_tree.column("ISBN", width=80)
        self.now_splite_has_books_show_tree.column("库存", width=40)
        self.now_splite_has_books_show_tree.column("id", width=120)
        self.now_splite_has_books_show_tree.place(x=10,y=50,width=780,height=400)
        self.now_splite_has_books_show_tree.bind("<ButtonRelease-1>", self.root_tree_click)

        self.now_splite_has_books_show_tree_refresh_button = maliang.Button(cv,(10,15),(50,30),fontsize=13,text="刷新",command=self.refresh_now_splite_has_books_show_tree)
        self.refresh_now_splite_has_books_show_tree()
        
        
        
        

        # 添加更多分类菜单
        # self.other_menu = tk.Menu(self.head_menus, tearoff=0)
        # self.other_menu.add_command(label="其他功能1", command=self.other_function1)
        # self.other_menu.add_command(label="其他功能2", command=self.other_function2)
        # self.head_menus.add_cascade(label="其他", menu=self.other_menu)

        
        def update_time():
            self.now_time.set(text=str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            # print(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            cv.after(1000, update_time)
        self.root.after(1000, update_time)
        # maliang.Button(cv, (20, 20), text="Button", command=lambda: print("Click"))
        # self.a =maliang.CheckBox(cv, (20, 20))
        # maliang.Text(cv, (60, 35), text="CheckBox", anchor="w")
        # maliang.CheckBox(cv, (20, 120), default=True)
        # maliang.Text(cv, (60, 135), text="CheckBox", anchor="w")
        # maliang.Button(cv, (20, 220), text="Button", command=lambda:self.p())

        self.root.protocol("WM_DELETE_WINDOW", lambda:exit())
        self.root.config(menu=self.head_menus)
        self.root.center()



        

        self.root.mainloop()
    
    def root_tree_click(self,event=None):
        pass
    def refresh_now_splite_has_books_show_tree(self):
        print("刷新")
        now_books_list = lb.Find_Books("")
        for item in self.now_splite_has_books_show_tree.get_children():
            self.now_splite_has_books_show_tree.delete(item)
        for book in now_books_list:
            self.now_splite_has_books_show_tree.insert("", "end", values=book)
        
    








    def return_book(self):
        print("单本书还书")
        self.root.withdraw()

        self.one_return_window = maliang.Toplevel(self.root,size=(1000,600),title="还书")
        self.one_return_window.center()
        self.one_return_window.iconbitmap(mypath("favicon.ico"))
        self.one_return_window__Canver = maliang.Canvas(self.one_return_window,auto_update=True,expand="xy",keep_ratio="max",auto_zoom=True)
        self.one_return_window__Canver.place(width=1000, height=600, x=0, y=0)
        def print_option_selected(index):
            self.one_return_search_type = index
            
            print(f"用户选择的操作类型: {("书籍搜索", "ISBN搜索","书籍条形码扫描")[index]},索引:{index}")
        self.one_return_back_botton = maliang.Button(self.one_return_window__Canver,(0,0),size=(50,20),fontsize=15,text="返回", anchor="nw", command=lambda:self.goback(self.one_return_window))
        self.one_return_search_type_text = maliang.Text(self.one_return_window__Canver,(0,90),text="选择操作类型", anchor="nw")
        self.one_return_search_type_OptionButton = maliang.OptionButton(self.one_return_window__Canver,(130,85), size=(140,40),text=("书籍搜索", "ISBN搜索","书籍条形码扫描"),command=print_option_selected,default=0)
        self.one_return_search_type_is_opencv_to_barcode_button = maliang.Button(self.one_return_window__Canver,(130,135),text="书籍条形码扫描", anchor="nw", command=lambda:self.opencv_for_book_isbn_barcode())
        self.one_return_is_teacher_or_student_text = maliang.Text(self.one_return_window__Canver,(0,200),text="还书人物:", anchor="nw")
        self.one_return_is_teacher_or_student_SegmentedButton = maliang.SegmentedButton(self.one_return_window__Canver,(120,195),sizes=((100,30),(100,30)),text=("学生","教师"),default=0)

        self.one_return_save_history_checkbox = maliang.CheckBox(self.one_return_window__Canver, (0, 260),default=True)
        self.one_return_save_history_text = maliang.Text(self.one_return_window__Canver,(45,260),text="保存学生还书历史", anchor="nw")

        self.one_return_ready_to_return_text = maliang.Text(self.one_return_window__Canver,(400,35),text="待还书书籍:", anchor="nw")
        self.one_return_ready_to_return_tree = ttk.Treeview(self.one_return_window__Canver,height=2,columns=("书名","作者","出版社","出版时间"),show="headings")
        self.one_return_ready_to_return_tree.heading("书名", text="书名")
        self.one_return_ready_to_return_tree.heading("作者", text="作者")
        self.one_return_ready_to_return_tree.heading("出版社", text="出版社")
        self.one_return_ready_to_return_tree.heading("出版时间", text="出版时间")
        self.one_return_ready_to_return_tree.column("书名", width=100)
        self.one_return_ready_to_return_tree.column("作者", width=100)
        self.one_return_ready_to_return_tree.column("出版社", width=100)
        self.one_return_ready_to_return_tree.column("出版时间", width=100)
        self.one_return_ready_to_return_tree.place(x=530,y=10,width=400,height=70)
        self.one_return_ready_to_return_tree.bind("<ButtonRelease-1>", self.ready_to_return_book_touch)

        self.one_return_search_button = maliang.Button(self.one_return_window__Canver,(400,115),size=(100,40),fontsize=15,text="搜索", anchor="nw", command=lambda:self.search_return_book())
        
        self.one_return_search_inputbox = maliang.InputBox(self.one_return_window__Canver,(500,115),size=(500,40))

        self.one_return_search_show_books_tree = ttk.Treeview(self.one_return_window__Canver,columns=("书名","作者","出版社","出版时间","ISBN","库存"),show="headings")
        self.one_return_search_show_books_tree.heading("书名", text="书名")
        self.one_return_search_show_books_tree.heading("作者", text="作者")
        self.one_return_search_show_books_tree.heading("出版社", text="出版社")
        self.one_return_search_show_books_tree.heading("出版时间", text="出版时间")
        self.one_return_search_show_books_tree.heading("ISBN", text="ISBN")
        self.one_return_search_show_books_tree.heading("库存", text="库存")
        self.one_return_search_show_books_tree.column("书名", width=100)
        self.one_return_search_show_books_tree.column("作者", width=100)
        self.one_return_search_show_books_tree.column("出版社", width=100)
        self.one_return_search_show_books_tree.column("出版时间", width=100)
        self.one_return_search_show_books_tree.column("ISBN", width=100)
        self.one_return_search_show_books_tree.column("库存", width=100)
        self.one_return_search_show_books_tree.place(x=400,y=160,width=600)
        self.one_return_search_show_books_tree.bind("<ButtonRelease-1>", self.search_touch_add_to_ready_to_return_tree)

        self.one_return_opencv_to_student_qrcode = maliang.Button(self.one_return_window__Canver,(0,330),size=(150,40),fontsize=15,text="扫描学生二维码", anchor="nw", command=lambda:self.qrcode_to_student_for_return())
        self.one_return_opencv_return_name = maliang.Text(self.one_return_window__Canver,(180,330),text="学生姓名:", anchor="nw")
        self.one_return_opencv_return_class = maliang.Text(self.one_return_window__Canver,(180,430),text="学生班级:", anchor="nw")
        self.one_return_opencv_return_id = maliang.Text(self.one_return_window__Canver,(180,380),text="学生座号:", anchor="nw")
        
        self.one_return_return_button = maliang.Button(self.one_return_window__Canver,(450,520),size=(100,40),fontsize=15,text="还书", anchor="nw", command=lambda:self.one_return_window_return_book_click())



        self.one_return_window.protocol("WM_DELETE_WINDOW", lambda:self.goback(self.one_return_window))
        pass

    def borrow_book(self):
        self.root.withdraw()

        self.one_borrow_window = maliang.Toplevel(self.root,size=(1000,600),title="借阅图书")
        self.one_borrow_window.center()
        self.one_borrow_window.iconbitmap(mypath("favicon.ico"))
        self.one_borrow_window__Canver = maliang.Canvas(self.one_borrow_window,auto_update=True,expand="xy",keep_ratio="max",auto_zoom=True)
        self.one_borrow_window__Canver.place(width=1000, height=600, x=0, y=0)
        #添加一个按钮在窗口左上角
        self.one_borrow_back_botton = maliang.Button(self.one_borrow_window__Canver,(0,0),size=(50,20),fontsize=15,text="返回", anchor="nw", command=lambda:self.goback(self.one_borrow_window))
        # self.one_borrow_window__Canver.create_line(500, 0, 500, 400, fill="blue violet")


        def print_option_selected(index):
            self.one_borrow_search_type = index
            print(f"用户选择的操作类型: {("书籍搜索", "ISBN搜索")[index]},索引:{index}")
        self.one_borrow_search_type_text = maliang.Text(self.one_borrow_window__Canver,(0,120),text="选择操作类型", anchor="nw")
        self.one_borrow_search_type_OptionButton = maliang.OptionButton(self.one_borrow_window__Canver,(130,115), size=(100,40),text=("书籍搜索", "ISBN搜索"),command=print_option_selected,default=0)

        self.one_borrow_is_teacher_or_student_text = maliang.Text(self.one_borrow_window__Canver,(0,200),text="借书人物:", anchor="nw")
        self.one_borrow_is_teacher_or_student_SegmentedButton = maliang.SegmentedButton(self.one_borrow_window__Canver,(120,195),sizes=((100,30),(100,30)),text=("学生","教师"),default=0)

        self.one_borrow_save_history_checkbox = maliang.CheckBox(self.one_borrow_window__Canver, (0, 260),default=True)
        self.one_borrow_save_history_text = maliang.Text(self.one_borrow_window__Canver,(45,260),text="保存学生借阅历史", anchor="nw")

        self.one_borrow_ready_to_borrow_text = maliang.Text(self.one_borrow_window__Canver,(400,35),text="待借阅书籍:", anchor="nw")
        self.one_borrow_ready_to_borrow_tree = ttk.Treeview(self.one_borrow_window__Canver,height=2,columns=("书名","作者","出版社","出版时间"),show="headings")
        self.one_borrow_ready_to_borrow_tree.heading("书名", text="书名")
        self.one_borrow_ready_to_borrow_tree.heading("作者", text="作者")
        self.one_borrow_ready_to_borrow_tree.heading("出版社", text="出版社")
        self.one_borrow_ready_to_borrow_tree.heading("出版时间", text="出版时间")
        self.one_borrow_ready_to_borrow_tree.column("书名", width=100)
        self.one_borrow_ready_to_borrow_tree.column("作者", width=100)
        self.one_borrow_ready_to_borrow_tree.column("出版社", width=100)
        self.one_borrow_ready_to_borrow_tree.column("出版时间", width=100)
        self.one_borrow_ready_to_borrow_tree.place(x=530,y=10,width=400,height=70)
        self.one_borrow_ready_to_borrow_tree.bind("<ButtonRelease-1>", self.ready_to_borrow_book_touch)

        self.one_borrow_search_button = maliang.Button(self.one_borrow_window__Canver,(400,115),size=(100,40),fontsize=15,text="搜索", anchor="nw", command=lambda:self.search_book())
        
        self.one_borrow_search_inputbox = maliang.InputBox(self.one_borrow_window__Canver,(500,115),size=(500,40))

        self.one_borrow_search_show_books_tree = ttk.Treeview(self.one_borrow_window__Canver,columns=("书名","作者","出版社","出版时间","ISBN","库存"),show="headings")
        self.one_borrow_search_show_books_tree.heading("书名", text="书名")
        self.one_borrow_search_show_books_tree.heading("作者", text="作者")
        self.one_borrow_search_show_books_tree.heading("出版社", text="出版社")
        self.one_borrow_search_show_books_tree.heading("出版时间", text="出版时间")
        self.one_borrow_search_show_books_tree.heading("ISBN", text="ISBN")
        self.one_borrow_search_show_books_tree.heading("库存", text="库存")
        self.one_borrow_search_show_books_tree.column("书名", width=100)
        self.one_borrow_search_show_books_tree.column("作者", width=100)
        self.one_borrow_search_show_books_tree.column("出版社", width=100)
        self.one_borrow_search_show_books_tree.column("出版时间", width=100)
        self.one_borrow_search_show_books_tree.column("ISBN", width=100)
        self.one_borrow_search_show_books_tree.column("库存", width=100)
        self.one_borrow_search_show_books_tree.place(x=400,y=160,width=600)
        self.one_borrow_search_show_books_tree.bind("<ButtonRelease-1>", self.search_touch_add_to_ready_to_borrow_tree)

        self.one_borrow_opencv_to_student_qrcode = maliang.Button(self.one_borrow_window__Canver,(0,330),size=(150,40),fontsize=15,text="扫描学生二维码", anchor="nw", command=lambda:self.qrcode_to_student())
        self.one_borrow_opencv_return_name = maliang.Text(self.one_borrow_window__Canver,(180,330),text="学生姓名:", anchor="nw")
        self.one_borrow_opencv_return_class = maliang.Text(self.one_borrow_window__Canver,(180,430),text="学生班级:", anchor="nw")
        self.one_borrow_opencv_return_id = maliang.Text(self.one_borrow_window__Canver,(180,380),text="学生座号:", anchor="nw")
        
        self.one_borrow_borrow_button = maliang.Button(self.one_borrow_window__Canver,(450,520),size=(100,40),fontsize=15,text="借书", anchor="nw", command=lambda:self.one_borrow_window_borrow_book_click())
        self.one_borrow_window.protocol("WM_DELETE_WINDOW", lambda:self.goback(self.one_borrow_window))

    def import_book(self):
        #先隐藏root窗口
        self.root.withdraw()
        
        self._import_book_ = maliang.Toplevel(self.root,size=(1000,600),title="导入图书")
        self._import_book_.center()
        self._import_book_.iconbitmap(mypath("favicon.ico"))
        self._import_book__Canver = maliang.Canvas(self._import_book_,auto_update=True,expand="xy",keep_ratio="max",auto_zoom=True)
        self._import_book__Canver.place(width=1000, height=600, x=0)
        #添加一个按钮在窗口左上角
        self._import_book_back_botton = maliang.Button(self._import_book__Canver,(0,0),size=(50,20),fontsize=15,text="返回", anchor="nw", command=lambda:self.goback(self._import_book_))
        # cv = maliang.Canvas(auto_zoom=True)
        # cv.place(width=1280, height=720)

        # maliang.Button(cv, (20, 20), text="Button", command=lambda: print("Click"))
        # maliang.CheckBox(self._import_book__Canver, (0, 80),length=30, command=print)
        # maliang.CheckBox(self._import_book__Canver, (30, 80), command=print)
        # maliang.Text(self._import_book__Canver, (60, 95), text="CheckBox", anchor="w")
        # maliang.CheckBox(self._import_book__Canver, (20, 120), command=print, default=True)
        # maliang.Text(self._import_book__Canver, (60, 135), text="CheckBox", anchor="w")
        # rb1 = maliang.RadioBox(self._import_book__Canver, (20, 20), command=print)
        # maliang.Text(self._import_book__Canver, (60, 35), text="", anchor="w")
        # maliang.RadioBox.group(rb1)


        #添加导入按钮
        self._import_book_import_button = maliang.Button(self._import_book__Canver,(0,30),size=(50,20),fontsize=20,text="上传", anchor="nw", command=lambda:self.import_excel())
        #添加刷新按钮
        self._import_book_import_button = maliang.Button(self._import_book__Canver,(0,60),size=(50,20),fontsize=20,text="刷新", anchor="nw", command=lambda:self.updata_treeview())
        #添加一个表格
        self._import_book_table = ttk.Treeview(self._import_book__Canver,height=2,columns=("书名","作者","出版社","出版时间","书籍介绍","ISBN","库存"),show="headings")
        self._import_book_table.heading("书名", text="书名")
        self._import_book_table.heading("作者", text="作者")
        self._import_book_table.heading("出版社", text="出版社")
        self._import_book_table.heading("出版时间", text="出版时间")
        self._import_book_table.heading("书籍介绍", text="书籍介绍")
        self._import_book_table.heading("库存", text="库存")
        self._import_book_table.heading("ISBN", text="ISBN")
        self._import_book_table.column("书名", width=100)
        self._import_book_table.column("作者", width=100)
        self._import_book_table.column("出版社", width=100)
        self._import_book_table.column("出版时间", width=100)
        self._import_book_table.column("书籍介绍", width=150)
        self._import_book_table.column("库存", width=50)
        self._import_book_table.column("ISBN", width=100)
        self._import_book_table.place(x=60,y=0,width=970,height=100)

        self._import_book_import_start_row_text = maliang.Text(self._import_book__Canver, position=(150, 110),text="书籍信息从excel表第")
        self._import_book_import_start_row_spinbox = maliang.SpinBox(self._import_book__Canver, position=(350, 110),size=(100, 30))
        self._import_book_import_start_row_text2 = maliang.Text(self._import_book__Canver, position=(455, 110),text="行开始")
        self._import_book_import_start_row_spinbox.set(self.excel_import_start_row)
        self._import_book_import_excel_column_text = maliang.Text(self._import_book__Canver, position=(850, 110))
        self._import_book_import_excel_column_text.set("表格 列:0")
        self._import_book_import_excel_row_text = maliang.Text(self._import_book__Canver, position=(850, 160))
        self._import_book_import_excel_row_text.set("表格 行:0")

        self._import_book_import_bookname_column_text = maliang.Text(self._import_book__Canver, position=(0, 150),text="书名 列为")
        self._import_book_import_bookname_column_spinbox = maliang.SpinBox(self._import_book__Canver, position=(100, 150),size=(100, 30))
        self._import_book_import_author_column_text = maliang.Text(self._import_book__Canver, position=(0, 190),text="作者 列为")
        self._import_book_import_author_column_spinbox = maliang.SpinBox(self._import_book__Canver, position=(100, 190),size=(100, 30))
        self._import_book_import_press_column_text = maliang.Text(self._import_book__Canver, position=(0, 230),text="出版社 列为")
        self._import_book_import_press_column_spinbox = maliang.SpinBox(self._import_book__Canver, position=(120, 230),size=(100, 30))
        self._import_book_import_publicationTime_column_text = maliang.Text(self._import_book__Canver, position=(0, 270),text="出版时间 列为")
        self._import_book_import_publicationTime_column_spinbox = maliang.SpinBox(self._import_book__Canver, position=(140, 270),size=(100, 30))
        self._import_book_import_bookInfo_column_text = maliang.Text(self._import_book__Canver, position=(0, 310),text="书籍介绍 列为")
        self._import_book_import_bookInfo_column_spinbox = maliang.SpinBox(self._import_book__Canver, position=(140, 310),size=(100, 30))
        self._import_book_import_isbn_column_text = maliang.Text(self._import_book__Canver, position=(0, 350),text="ISBN 列为")
        self._import_book_import_isbn_column_spinbox = maliang.SpinBox(self._import_book__Canver, position=(100, 350),size=(100, 30))
        self._import_book_import_inventory_column_text = maliang.Text(self._import_book__Canver, position=(0, 390),text="库存 列为")
        self._import_book_import_inventory_column_spinbox = maliang.SpinBox(self._import_book__Canver, position=(100, 390),size=(100, 30))

        self._import_book_continue_the_None_in_excel_every_checkbox = maliang.CheckBox(self._import_book__Canver, position=(430, 400))
        self._import_book_continue_the_None_in_excel_every_text = maliang.Text(self._import_book__Canver, position=(470, 400),text="是否跳过空行")

        self._import_book_import_bookname_column_spinbox.set(self.excel_import_bookname_column)
        self._import_book_import_author_column_spinbox.set(self.excel_import_author_column)
        self._import_book_import_press_column_spinbox.set(self.excel_import_press_column)
        self._import_book_import_publicationTime_column_spinbox.set(self.excel_import_publicationTime_column)
        self._import_book_import_bookInfo_column_spinbox.set(self.excel_import_bookInfo_column)
        self._import_book_import_isbn_column_spinbox.set(self.excel_import_isbn_column)
        self._import_book_import_inventory_column_spinbox.set(self.excel_import_inventory_column)

        self._import_book_import_to_splite_button = maliang.Button(self._import_book__Canver,(450,480),fontsize=20,text="导入至数据库", anchor="nw", command=lambda:self.import_to_splite())

        self.import_excel_start_row_change()
        print("导入书籍库")
        #如果关闭窗口则显示root窗口
        self._import_book_.protocol("WM_DELETE_WINDOW", lambda:self.goback(self._import_book_))
        pass
    
    def goback(self,thing):
        thing.destroy()
        self.root.deiconify()
        self.refresh_now_splite_has_books_show_tree()

    

    def import_excel(self,file_path:str=None,import_or_updata:float=True,import_to_splte:float=False):
        if import_or_updata:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            self.file_path = file_path
        print(self.file_path)
        
        if self.file_path:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            if self.check_user_set_import_about_excel_column_and_row(ws.max_column,ws.max_row):
                self.updata_to_init_one_borrow()
                if import_to_splte == False:
                    for thing in self._import_book_table.get_children():
                        self._import_book_table.delete(thing)
                    for i in range(int(self.excel_import_start_row),int(self.excel_import_start_row)+2):
                        A = str(ws.cell(column=int(self.excel_import_bookname_column),row=i).value)
                        B = str(ws.cell(column=int(self.excel_import_author_column),row=i).value)
                        C = str(ws.cell(column=int(self.excel_import_press_column),row=i).value)
                        D = str(ws.cell(column=int(self.excel_import_publicationTime_column),row=i).value)
                        E = str(ws.cell(column=int(self.excel_import_bookInfo_column),row=i).value)
                        F = str(ws.cell(column=int(self.excel_import_isbn_column),row=i).value)
                        G = str(ws.cell(column=int(self.excel_import_inventory_column),row=i).value)
                        book = (A,B,C,D,E,F,G)
                        print(book)
                        print(self._import_book_continue_the_None_in_excel_every_checkbox.get())
                        if self._import_book_continue_the_None_in_excel_every_checkbox.get() == False:
                            self._import_book_table.insert("", "end", values=book)
                        else:
                            if A == 'None' or B == 'None' or C == 'None' or D == 'None' or E == 'None' or F == 'None' or G == 'None':
                                print("跳过空行")
                                self._import_book_table.insert("", "end", values=("跳过","跳过","跳过","跳过","跳过","跳过","跳过"))
                                continue
                            else:
                                self._import_book_table.insert("", "end", values=book)
                    self._import_book_import_excel_column_text.set(f"表格 列:{ws.max_column}")
                    self._import_book_import_excel_row_text.set(f"表格 行:{ws.max_row}")
                    return 0
                else:
                    
                    self.jiazai = maliang.Toplevel(self._import_book_,size=(200,200),title="导入中")
                    self.jiazai.center()
                    self.jiazai.iconbitmap(mypath("favicon.ico"))
                    self.jiazai_Canver = maliang.Canvas(self.jiazai,auto_update=True,expand="xy",keep_ratio="max",auto_zoom=True)
                    self.jiazai_Canver.place(x=0,y=0,width=200,height=200)
                    self.jiazai_Spinner = maliang.Spinner(self.jiazai_Canver,position=(0,0),size=(200,200),auto_update=True)
                    self.jiazai_jindu_text = maliang.Text(self.jiazai_Canver,position=(40,85),text="进度: ")
                    all = ws.max_row+1 - int(self.excel_import_start_row)
                    global start 
                    start = int(self.excel_import_start_row)
                    end = ws.max_row
                    def imp(i):
                        A = str(ws.cell(column=int(self.excel_import_bookname_column),row=i).value)
                        B = str(ws.cell(column=int(self.excel_import_author_column),row=i).value)
                        C = str(ws.cell(column=int(self.excel_import_press_column),row=i).value)
                        D = str(ws.cell(column=int(self.excel_import_publicationTime_column),row=i).value)
                        E = str(ws.cell(column=int(self.excel_import_bookInfo_column),row=i).value)
                        F = str(ws.cell(column=int(self.excel_import_isbn_column),row=i).value)
                        G = str(ws.cell(column=int(self.excel_import_inventory_column),row=i).value)
                        _to_add_book_list = [A,B,C,D,E,F,G]
                        continue_None_in_excel_every = self._import_book_continue_the_None_in_excel_every_checkbox.get()
                        if continue_None_in_excel_every == False:
                            lb.Add_Book(_to_add_book_list)
                        else:
                            None_num = 0
                            for _to_add_book_every_message in _to_add_book_list:
                                if _to_add_book_every_message == 'None':
                                    None_num += 1
                                    break
                            if None_num == 0:
                                lb.Add_Book(_to_add_book_list)
                    def jiazaiing():
                        global start
                        imp(start)
                        self.jiazai_Spinner.set(round(start/all,2))
                        self.jiazai_jindu_text.set(f"进度: {round(start/all,2)*100}%")
                        if start == end:
                            messagebox.showinfo("提示","导入完成")
                            self.goback(self.jiazai)
                            self.goback(self._import_book_)
                            return 0
                        else:
                            start += 1
                            self.jiazai.after(10,jiazaiing)
                    self.jiazai.after(100,jiazaiing)
                    self.jiazai.mainloop()
                    
            else:
                print("表格列数输入有误")
                return 0
        else:
            print("未选择文件")
            return 0
                
    def check_user_set_import_about_excel_column_and_row(self,max_column,max_row):
        show_message_to_user = ""
        user_choise_start_row = self._import_book_import_start_row_spinbox.get()
        try:
            if int(user_choise_start_row) > max_row or int(user_choise_start_row) < 1:
                show_message_to_user += f"Excel表格行数不足,最大{max_row},最小1，请重新选择起始行数\n"
            user_choise_bookname_column = int(self._import_book_import_bookname_column_spinbox.get())
            user_choise_author_column = int(self._import_book_import_author_column_spinbox.get())
            user_choise_press_column = int(self._import_book_import_press_column_spinbox.get())
            user_choise_publicationTime_column = int(self._import_book_import_publicationTime_column_spinbox.get())
            user_choise_bookInfo_column = int(self._import_book_import_bookInfo_column_spinbox.get())
            user_choise_isbn_column = int(self._import_book_import_isbn_column_spinbox.get())
            user_choise_inventory_column = int(self._import_book_import_inventory_column_spinbox.get())
        except ValueError:
            messagebox.showerror("错误","表格列或开始行数输入有误,请重新输入\n")
            return False
        user_chiose_bookmessage_list = [user_choise_bookname_column,user_choise_author_column,user_choise_press_column,user_choise_publicationTime_column,user_choise_bookInfo_column,user_choise_isbn_column,user_choise_inventory_column]
        #检查user_chiose_bookmessage_list是否有重复的数字
        if len(user_chiose_bookmessage_list) != len(set(user_chiose_bookmessage_list)):
            show_message_to_user += "表格列数不能重复,请重新选择列数\n"
        for i in user_chiose_bookmessage_list:
            if i > max_column or i < 1:
                show_message_to_user += f"表格列数有误,最大{max_column}，最小1，请重新选择列数\n"
        if len(show_message_to_user) >0:
            messagebox.showerror("错误",show_message_to_user)
            return False
        else:
            return True
    def updata_to_init_one_borrow(self):
        _start_row =self._import_book_import_start_row_spinbox.get()
        self.excel_import_start_row = _start_row
        _bookname_column =self._import_book_import_bookname_column_spinbox.get()
        self.excel_import_bookname_column = _bookname_column
        _author_column =self._import_book_import_author_column_spinbox.get()
        self.excel_import_author_column = _author_column
        _press_column =self._import_book_import_press_column_spinbox.get()
        self.excel_import_press_column = _press_column
        _publicationTime_column =self._import_book_import_publicationTime_column_spinbox.get()
        self.excel_import_publicationTime_column = _publicationTime_column
        _bookInfo_column =self._import_book_import_bookInfo_column_spinbox.get()
        self.excel_import_bookInfo_column = _bookInfo_column
        _isbn_column =self._import_book_import_isbn_column_spinbox.get()
        self.excel_import_isbn_column = _isbn_column
        _inventory_column =self._import_book_import_inventory_column_spinbox.get()
        self.excel_import_inventory_column = _inventory_column
    def import_excel_start_row_change(self):
        excel_import_start_row = self._import_book_import_start_row_spinbox.get()
        print(excel_import_start_row)
        self.excel_import_start_row = excel_import_start_row
        print(self.excel_import_start_row)
    def updata_treeview(self):
        self.import_excel(import_or_updata=False)
        pass
    def import_to_splite(self):
        if self.file_path != None:
            self.import_excel(import_or_updata=False,import_to_splte=True)
        else:
            print("未选择文件")
        return 0
    
    def delete_all_book(self):
        lb.delete_all_book()
        self.refresh_now_splite_has_books_show_tree()
        messagebox.showinfo("提示","已清空所有书籍信息")
        return 0
    
    def search_book(self):
        if self.one_borrow_search_type_OptionButton.get() != None:
            if self.one_borrow_search_type == 0:
                #书内容搜索
                search_result = lb.Find_Books(self.one_borrow_search_inputbox.get())
                print(search_result)
                for want_to_delete_book in self.one_borrow_search_show_books_tree.get_children():
                    self.one_borrow_search_show_books_tree.delete(want_to_delete_book)
                self.now_search_borrow_books_list = search_result
                for add_to_tree in search_result:
                    print(add_to_tree)
                    self.one_borrow_search_show_books_tree.insert("", "end", values=(add_to_tree[0],add_to_tree[1],add_to_tree[2],add_to_tree[3],add_to_tree[5],add_to_tree[6]))
                
            elif self.one_borrow_search_type == 1:
                #ISBN搜索
                search_result = lb.Find_book_by_isbn(self.one_borrow_search_inputbox.get())
                for want_to_delete_book in self.one_borrow_search_show_books_tree.get_children():
                    self.one_borrow_search_show_books_tree.delete(want_to_delete_book)
                if search_result['code'] == 200:
                    add_to_tree = search_result['msg']
                    print(add_to_tree)
                    self.now_search_borrow_books_list = [add_to_tree]
                    self.one_borrow_search_show_books_tree.insert("", "end", values=(add_to_tree[0],add_to_tree[1],add_to_tree[2],add_to_tree[3],add_to_tree[5],add_to_tree[6]))
                pass
        

    def search_touch_add_to_ready_to_borrow_tree(self,event):
        selected_item = self.one_borrow_search_show_books_tree.selection()[0]
        book_info_in_tree = self.one_borrow_search_show_books_tree.item(selected_item, "values")
        index = self.one_borrow_search_show_books_tree.index(selected_item)
        book_info = self.now_search_borrow_books_list[index]
        print(book_info)
        book_name = book_info[0]
        book_author = book_info[1]
        book_press = book_info[2]
        book_publicationTime = book_info[3]
        book_isbn = book_info[5]
        if book_info_in_tree not in self.want_to_borrow_books_list:
            self.want_to_borrow_books_list.append(book_info_in_tree)
            self.want_to_borrow_books_list_name_and_isbn_tuple.append((book_name, book_isbn))
            self.one_borrow_ready_to_borrow_tree.insert("", "end", values=(book_name, book_author, book_press, book_publicationTime))
            messagebox.showinfo("提示", "添加成功")
            

    def ready_to_borrow_book_touch(self,event):
        selected_item = self.one_borrow_ready_to_borrow_tree.selection()[0]
        selected_index = self.one_borrow_ready_to_borrow_tree.index(selected_item)
        self.want_to_borrow_books_list.pop(selected_index)
        for m in self.want_to_borrow_books_list_name_and_isbn_tuple:
            if m[0] == self.one_borrow_ready_to_borrow_tree.item(selected_item, "values")[0]:
                self.want_to_borrow_books_list_name_and_isbn_tuple.remove(m)
                break
        self.one_borrow_ready_to_borrow_tree.delete(selected_item)
        # messagebox.showinfo("提示", "删除成功")

    def one_borrow_window_borrow_book_click(self):
        if self.one_borrow_is_teacher_or_student_SegmentedButton.get() == 0:
            if self.borrow_student_name!=""or self.borrow_student_id!=""or self.borrow_student_class!=""or self.borrow_student_password!=""or self.borrow_student_borrow_books!=""or self.borrow_student_borrowed_books!="":
                print("================================================================================")
                print(f"{("学生","老师")[self.one_borrow_is_teacher_or_student_SegmentedButton.get()]} 借书")
                print(f"是否保留借书历史: {self.one_borrow_save_history_checkbox.get()}")
                print(f"学生姓名: {self.borrow_student_name}")
                print(f"学生座号: {self.borrow_student_id}")
                print(f"学生班级: {self.borrow_student_class}")
                print(f"学生借书书籍: {self.borrow_student_borrow_books}")
                print(f"学生借书密码: {self.borrow_student_password}")
                print(f"学生借书历史: {self.borrow_student_borrowed_books}")
                #print(f"借书书籍: {self.want_to_borrow_books_list}")
                show_to_user = ""
                if len(self.want_to_borrow_books_list) > 0:
                    for to_borrow_isbn in self.want_to_borrow_books_list_name_and_isbn_tuple:
                        print(to_borrow_isbn)
                        print(to_borrow_isbn[1])
                        #lb.borrow_book(toborrow_book_isbn,list[name,id,class])
                        show_to_user +=f"{to_borrow_isbn[0]} {lb.Borrow_Book(to_borrow_isbn[1],[self.borrow_student_name,self.borrow_student_id,self.borrow_student_class,self.borrow_student_borrow_books,self.borrow_student_password,self.borrow_student_borrowed_books],save_history=self.one_borrow_save_history_checkbox.get())} \n"
                else:
                    messagebox.showerror("错误","请先选择要借的书籍")
                    return 0
            else:
                messagebox.showerror("错误","请先添加学生信息")
        elif self.one_borrow_is_teacher_or_student_SegmentedButton.get() == 1:
            show_to_user = ""
            if len(self.want_to_borrow_books_list) > 0:
                for to_borrow_isbn in self.want_to_borrow_books_list_name_and_isbn_tuple:
                    print(to_borrow_isbn)
                    print(to_borrow_isbn[1])
                    show_to_user +=f"{to_borrow_isbn[0]} {lb.Borrow_Book(to_borrow_isbn[1],[],save_history=False)} \n"
            else:
                messagebox.showerror("错误","请先选择要借的书籍")
                return 0
        messagebox.showinfo("提示", show_to_user)
        for it in self.one_borrow_ready_to_borrow_tree.get_children():
            self.one_borrow_ready_to_borrow_tree.delete(it)
        self.want_to_borrow_books_list = []
        self.want_to_borrow_books_list_name_and_isbn_tuple = []
        self.borrow_student_name = ""
        self.borrow_student_id = ""
        self.borrow_student_class = ""
        self.borrow_student_password = ""
        self.borrow_student_borrow_books = ""
        self.borrow_student_borrowed_books = ""
        self.one_borrow_opencv_return_name.set(text=f"学生姓名: ")
        self.one_borrow_opencv_return_id.set(text=f"学生座号: ")
        self.one_borrow_opencv_return_class.set(text=f"学生班级: ")
        self.goback(self.one_borrow_window)


    def qrcode_to_student(self):
        if self.one_borrow_is_teacher_or_student_SegmentedButton.get() == 0:
            student_msg = lb.cv_for_student()
            jianli_lianjie()
            student_msg = lb.Login_User(student_msg[0], student_msg[2], student_msg[1], student_msg[3])
            print(student_msg)
            if student_msg['code'] == 200:
                student_msg = student_msg['msg'][0]
                student_name = student_msg[0]
                self.borrow_student_name = student_name
                student_id = student_msg[1]
                self.borrow_student_id = student_id
                student_class = student_msg[2]
                self.borrow_student_class = student_class
                student_borrow_book = student_msg[3]
                self.borrow_student_borrow_books = student_borrow_book
                student_borrow_password = student_msg[4]
                self.borrow_student_password = student_borrow_password
                student_borrow_history = student_msg[5]
                self.borrow_student_borrowed_books = student_borrow_history
                self.one_borrow_opencv_return_name.set(text=f"学生姓名: {student_name}")
                self.one_borrow_opencv_return_id.set(text=f"学生座号: {student_id}")
                self.one_borrow_opencv_return_class.set(text=f"学生班级: {student_class}")
                
            elif student_msg['code'] == 404:
                messagebox.showinfo("提示", "未找到该学生")
                return 0



    def ready_to_return_book_touch(self,event):
        """selected_item = self.one_borrow_ready_to_borrow_tree.selection()[0]
        selected_index = self.one_borrow_ready_to_borrow_tree.index(selected_item)
        self.want_to_borrow_books_list.pop(selected_index)
        for m in self.want_to_borrow_books_list_name_and_isbn_tuple:
            if m[0] == self.one_borrow_ready_to_borrow_tree.item(selected_item, "values")[0]:
                self.want_to_borrow_books_list_name_and_isbn_tuple.remove(m)
                break
        self.one_borrow_ready_to_borrow_tree.delete(selected_item)
        # messagebox.showinfo("提示", "删除成功")
        """
        selected_item = self.one_return_ready_to_return_tree.selection()[0]
        selected_index = self.one_return_ready_to_return_tree.index(selected_item)
        self.want_to_return_books_list.pop(selected_index)
        for m in self.want_to_return_books_list_name_and_isbn_tuple:
            if m[0] == self.one_return_ready_to_return_tree.item(selected_item, "values")[0]:
                self.want_to_return_books_list_name_and_isbn_tuple.remove(m)
                break
        self.one_return_ready_to_return_tree.delete(selected_item)
        # messagebox.showinfo("提示", "删除成功")

    def search_return_book(self):
        print(self.one_return_search_type_OptionButton.get())
        if self.one_return_search_type_OptionButton.get() != None and self.one_return_search_inputbox.get() != 2:
            if self.one_return_search_type_OptionButton.get() == 0:
                #书内容搜索
                search_result = lb.Find_Books(self.one_return_search_inputbox.get())
                print(search_result)
                for want_to_delete_book in self.one_return_search_show_books_tree.get_children():
                    self.one_return_search_show_books_tree.delete(want_to_delete_book)
                self.now_search_return_books_list = search_result
                for add_to_tree in search_result:
                    print(add_to_tree)
                    self.one_return_search_show_books_tree.insert("", "end", values=(add_to_tree[0],add_to_tree[1],add_to_tree[2],add_to_tree[3],add_to_tree[5],add_to_tree[6]))
                
            elif self.one_return_search_type_OptionButton.get() == 1:
                #ISBN搜索
                search_result = lb.Find_book_by_isbn(self.one_return_search_inputbox.get())
                for want_to_delete_book in self.one_return_search_show_books_tree.get_children():
                    self.one_return_search_show_books_tree.delete(want_to_delete_book)
                if search_result['code'] == 200:
                    add_to_tree = search_result['msg']
                    print(add_to_tree)
                    self.now_search_return_books_list = [add_to_tree]
                    self.one_return_search_show_books_tree.insert("", "end", values=(add_to_tree[0],add_to_tree[1],add_to_tree[2],add_to_tree[3],add_to_tree[5],add_to_tree[6]))
                elif search_result['code'] == 404:
                    messagebox.showinfo("提示", "未找到该书籍")
            

            

    def search_touch_add_to_ready_to_return_tree(self,event):
        selected_item = self.one_return_search_show_books_tree.selection()[0]
        book_info_in_tree = self.one_return_search_show_books_tree.item(selected_item, "values")
        index = self.one_return_search_show_books_tree.index(selected_item)
        book_info = self.now_search_return_books_list[index]
        print(book_info)
        book_name = book_info[0]
        book_author = book_info[1]
        book_press = book_info[2]
        book_publicationTime = book_info[3]
        book_isbn = book_info[5]
        if book_info_in_tree not in self.want_to_return_books_list:
            self.want_to_return_books_list.append(book_info_in_tree)
            self.want_to_return_books_list_name_and_isbn_tuple.append((book_name, book_isbn))
            self.one_return_ready_to_return_tree.insert("", "end", values=(book_name, book_author, book_press, book_publicationTime))
            messagebox.showinfo("提示", "添加成功")
        pass

    def qrcode_to_student_for_return(self):
        if self.one_return_is_teacher_or_student_SegmentedButton.get() == 0:
            student_msg = lb.cv_for_student()
            jianli_lianjie()
            student_msg = lb.Login_User(student_msg[0], student_msg[2], student_msg[1], student_msg[3])
            print(student_msg)
            if student_msg['code'] == 200:
                student_msg = student_msg['msg'][0]
                student_name = student_msg[0]
                self.return_student_name = student_name
                student_id = student_msg[1]
                self.return_student_id = student_id
                student_class = student_msg[2]
                self.return_student_class = student_class
                student_return_book = student_msg[3]
                self.return_student_borrow_books = student_return_book
                student_return_password = student_msg[4]
                self.return_student_password = student_return_password
                student_borrow_history = student_msg[5]
                self.return_student_borrowed_books = student_borrow_history
                self.one_return_opencv_return_name.set(text=f"学生姓名: {student_name}")
                self.one_return_opencv_return_id.set(text=f"学生座号: {student_id}")
                self.one_return_opencv_return_class.set(text=f"学生班级: {student_class}")
                
            elif student_msg['code'] == 404:
                messagebox.showinfo("提示", "未找到该学生")
                return 0

    def one_return_window_return_book_click(self):
        if self.one_return_is_teacher_or_student_SegmentedButton.get() == 0:
            #student
            if self.return_student_name!=""or self.return_student_id!=""or self.return_student_class!=""or self.return_student_password!=""or self.return_student_borrow_books!=""or self.return_student_borrowed_books!="":
                print("================================================================================")
                print(f"{("学生","老师")[self.one_return_is_teacher_or_student_SegmentedButton.get()]} 借书")
                print(f"是否保留借书历史: {self.one_return_save_history_checkbox.get()}")
                print(f"学生姓名: {self.return_student_name}")
                print(f"学生座号: {self.return_student_id}")
                print(f"学生班级: {self.return_student_class}")
                print(f"学生还书籍: {self.return_student_borrow_books}")
                print(f"学生还密码: {self.return_student_password}")
                print(f"学生还历史: {self.return_student_borrowed_books}")
                
                show_to_user = ""
                if len(self.want_to_return_books_list) > 0:
                    for to_return_isbn in self.want_to_return_books_list_name_and_isbn_tuple:
                        print(to_return_isbn)
                        print(to_return_isbn[1])
                        
                        if lb.Login_User_Has_Book([self.return_student_name,self.return_student_id,self.return_student_class,self.return_student_borrow_books,self.return_student_password,self.return_student_borrowed_books],str(to_return_isbn[1])):
                            show_to_user +=f"{to_return_isbn[0]} {lb.Return_Book(to_return_isbn[1],[self.return_student_name,self.return_student_id,self.return_student_class,self.return_student_borrow_books,self.return_student_password,self.return_student_borrowed_books],save_history=self.one_return_save_history_checkbox.get())} \n"
                        else:
                            show_to_user += f"{to_return_isbn[0]} 你未借过这本书 \n"
                            

                    messagebox.showinfo("提示", show_to_user)
                else:
                    messagebox.showerror("错误","请先选择要借的书籍")
                    return 0
            else:
                messagebox.showerror("错误","请先添加学生信息")
                return 0
        elif self.one_return_is_teacher_or_student_SegmentedButton.get() == 1:
            show_to_user = ""
            if len(self.want_to_return_books_list) > 0:
                for to_return_isbn in self.want_to_return_books_list_name_and_isbn_tuple:
                    print(to_return_isbn)
                    print(to_return_isbn[1])
                    show_to_user +=f"{to_return_isbn[0]} {lb.Return_Book(to_return_isbn[1],[],save_history=self.one_return_save_history_checkbox.get())} \n"
                
                messagebox.showinfo("提示", show_to_user)
                return 0
            else:
                messagebox.showerror("错误","请先选择要借的书籍")
                return 0
        
        for item in self.one_return_ready_to_return_tree.get_children():
            self.one_return_ready_to_return_tree.delete(item)
        self.one_return_search_type = None
        self.now_search_return_books_list = []
        self.want_to_return_books_list = []
        self.want_to_return_books_list_name_and_isbn_tuple = []
        self.return_student_name = ""
        self.return_student_id = ""
        self.return_student_class = ""
        self.return_student_password = ""
        self.return_student_borrow_books = ""
        self.return_student_borrowed_books = ""
        self.want_to_return_books_list = []
        self.goback(self.one_return_window)
    def opencv_for_book_isbn_barcode(self):
        if self.one_return_search_type_OptionButton.get() == 2:
            print("条件满足，OpenCV for isbn条形码")
            book_msg = lb.cv_for_book()
            jianli_lianjie()
            print(book_msg)
            if book_msg['code']==200:
                book_msg = book_msg['msg']
                print(book_msg)
                for item in self.one_return_search_show_books_tree.get_children():
                    self.one_return_search_show_books_tree.delete(item)
                self.now_search_return_books_list = []
                self.one_return_search_show_books_tree.insert("", "end", values=(book_msg[0],book_msg[1],book_msg[2],book_msg[3],book_msg[5],book_msg[6]))
                self.now_search_return_books_list.append(book_msg)
                self.one_return_search_inputbox.set("")
            elif book_msg['code']==404:
                messagebox.showinfo("提示", "未识别到该书籍，请检查是否破损或错误数据，请重新添加")
                return 0

            

    def delete_one_book(self):
        self.root.withdraw()
        self.one_delete_book_window = maliang.Toplevel(self.root,size=(1000,800),title="单书删除")
        self.one_delete_book_window.center()
        self.one_delete_book_window.iconbitmap(mypath("favicon.ico"))
        

        self.one_delete_book_window__Canvas = maliang.Canvas(self.one_delete_book_window,auto_update=True,expand="xy",keep_ratio="max",auto_zoom=True)
        self.one_delete_book_window__Canvas.place(width=1000, height=800, x=0, y=0)
        def print_option_selected(index):
            print(f"{['书籍搜索',"isbn搜索"][index]} {index}")

        
        self.one_delete_book_back = maliang.Button(self.one_delete_book_window__Canvas,(0,0),size=(50,20),fontsize=15,text="返回", anchor="nw", command=lambda:self.goback(self.one_delete_book_window))
        self.one_delete_book_search_type_OptionButton = maliang.OptionButton(self.one_delete_book_window__Canvas,(0,50), size=(100,35),fontsize=15,text=("书籍搜索", "ISBN搜索"),command=print_option_selected,default=0)
        self.one_delete_book_search_button = maliang.Button(self.one_delete_book_window__Canvas,(100,50),size=(100,35),fontsize=15,text="搜索书籍", anchor="nw", command=lambda:self.one_delete_book_search_book())
        self.one_delete_book_search_inputbox = maliang.InputBox(self.one_delete_book_window__Canvas,(200,50),size=(800,35))
        self.one_delete_book_search_inputbox.bind("<Return>",self.one_delete_book_search_book)
        self.one_delete_book_search_show_tree = ttk.Treeview(self.one_delete_book_window__Canvas,columns=("书名","作者","出版社","出版时间","ISBN","库存"),show="headings")
        self.one_delete_book_search_show_tree.heading("书名", text="书名")
        self.one_delete_book_search_show_tree.heading("作者", text="作者")
        self.one_delete_book_search_show_tree.heading("出版社", text="出版社")
        self.one_delete_book_search_show_tree.heading("出版时间", text="出版时间")
        self.one_delete_book_search_show_tree.heading("ISBN", text="ISBN")
        self.one_delete_book_search_show_tree.heading("库存", text="库存")
        self.one_delete_book_search_show_tree.column("书名", width=100)
        self.one_delete_book_search_show_tree.column("作者", width=100)
        self.one_delete_book_search_show_tree.column("出版社", width=100)
        self.one_delete_book_search_show_tree.column("出版时间", width=100)
        self.one_delete_book_search_show_tree.column("ISBN", width=100)
        self.one_delete_book_search_show_tree.column("库存", width=100)
        self.one_delete_book_search_show_tree.place(x=0,y=100,width=600,height=500)
        self.one_delete_book_search_show_tree.bind("<ButtonRelease-1>",self.one_delete_book_search_show_tree_double_click)

        self.one_delete_book_delete_text = maliang.Text(self.one_delete_book_window__Canvas,(630,100),text="预删除书籍信息:",underline=True,anchor="nw")
        self.one_delete_book_book_name_text = maliang.Text(self.one_delete_book_window__Canvas,(630,150),text="书名: ",anchor="nw")
        self.one_delete_book_book_name = maliang.Text(self.one_delete_book_window__Canvas,(680,150),text="",anchor="nw")
        self.one_delete_book_author_text = maliang.Text(self.one_delete_book_window__Canvas,(630,200),text="作者: ",anchor="nw")
        self.one_delete_book_author_name = maliang.Text(self.one_delete_book_window__Canvas,(680,200),text="",anchor="nw")
        self.one_delete_book_press_text = maliang.Text(self.one_delete_book_window__Canvas,(630,250),text="出版社: ",anchor="nw")
        self.one_delete_book_press = maliang.Text(self.one_delete_book_window__Canvas,(700,250),text="",anchor="nw")
        self.one_delete_book_publicationTime_text = maliang.Text(self.one_delete_book_window__Canvas,(630,300),text="出版时间: ",anchor="nw")
        self.one_delete_book_publicationTime = maliang.Text(self.one_delete_book_window__Canvas,(720,300),text="",anchor="nw")
        self.one_delete_book_bookInfo_text = maliang.Text(self.one_delete_book_window__Canvas,(630,350),text="书籍介绍: ",anchor="nw")
        self.one_delete_book_bookInfo = maliang.Text(self.one_delete_book_window__Canvas,(630,380),text="",anchor="nw")
        self.one_delete_book_isbn_text = maliang.Text(self.one_delete_book_window__Canvas,(630,500),text="ISBN: ",anchor="nw")
        self.one_delete_book_isbn = maliang.Text(self.one_delete_book_window__Canvas,(685,500),text="",anchor="nw")
        self.one_delete_book_inventory_text = maliang.Text(self.one_delete_book_window__Canvas,(630,550),text="库存: ",anchor="nw")
        self.one_delete_book_inventory = maliang.Text(self.one_delete_book_window__Canvas,(680,550),text="",anchor="nw")
        self.one_delete_book_id_text = maliang.Text(self.one_delete_book_window__Canvas,(630,600),text="id: ",anchor="nw")
        self.one_delete_book_id = maliang.Text(self.one_delete_book_window__Canvas,(660,600),text="",anchor="nw")

        self.one_delete_book_delete_button = maliang.Button(self.one_delete_book_window__Canvas,(470,700),size=(100,35),fontsize=15,text="删除书籍", anchor="nw", command=lambda:self.one_delete_book_delete_book_button_click())


        self.one_delete_book_window.protocol("WM_DELETE_WINDOW", lambda:self.goback(self.one_delete_book_window))
        pass








    def one_delete_book_search_book(self,event=None):
        search_thing = self.one_delete_book_search_inputbox.get()
        print(search_thing)
        search_books_list = []
        if self.one_delete_book_search_type_OptionButton.get() == 0:
            #书籍搜索
            search_books_list = lb.Find_Books(self.one_delete_book_search_inputbox.get())
            # print(search_books_list)
            if len(search_books_list)>0:
                self.one_delete_book_now_show_books_list = []
                for item in self.one_delete_book_search_show_tree.get_children():
                    self.one_delete_book_search_show_tree.delete(item)
                for item in search_books_list:
                    self.one_delete_book_now_show_books_list.append(item)
                    self.one_delete_book_search_show_tree.insert("", "end", values=(item[0],item[1],item[2],item[3],item[5],item[6]))
                    
            else:
                messagebox.showerror("错误", "未找到相关书籍")
            pass

        elif self.one_delete_book_search_type_OptionButton.get() == 1:
            #isbn搜索
            search_books_dict = lb.Find_book_by_isbn(self.one_delete_book_search_inputbox.get())
            if search_books_dict['code'] == 200:
                search_books_list = search_books_dict['msg']
                self.one_delete_book_now_show_books_list = []
                for item in self.one_delete_book_search_show_tree.get_children():
                    self.one_delete_book_search_show_tree.delete(item)
                self.one_delete_book_now_show_books_list.append(search_books_list)
                print(search_books_list)
                self.one_delete_book_search_show_tree.insert("", "end", values=(search_books_list[0],search_books_list[1],search_books_list[2],search_books_list[3],search_books_list[5],search_books_list[6]))
            elif search_books_dict['code'] == 404:
                messagebox.showerror("错误", "未找到该书籍")
                return 0
            
        

    def one_delete_book_search_show_tree_double_click(self,event):
        selected_item = self.one_delete_book_search_show_tree.selection()[0]
        book_info_in_tree = self.one_delete_book_search_show_tree.item(selected_item, "values")
        index = self.one_delete_book_search_show_tree.index(selected_item)
        book_info = self.one_delete_book_now_show_books_list[index]
        print(book_info)
        #__init__里的
        self._one_delete_book_book_name = book_info[0]
        self._one_delete_book_author_name = book_info[1]
        self._one_delete_book_press = book_info[2]
        self._one_delete_book_publicationTime = book_info[3]
        self._one_delete_book_bookInfo = book_info[4]
        self._one_delete_book_isbn = book_info[5]
        self._one_delete_book_inventory = book_info[6]
        self._one_delete_book_id = book_info[7]
        
        self.one_delete_book_book_name.set(self._one_delete_book_book_name)
        self.one_delete_book_author_name.set(self._one_delete_book_author_name)
        self.one_delete_book_press.set(self._one_delete_book_press)
        self._one_delete_book_book_Info = self._one_delete_book_bookInfo.replace("\n", "")
        self._one_delete_book_book_Info = self._one_delete_book_book_Info.replace("\"", "'")
        print(self._one_delete_book_book_Info)
        info_len = len(self._one_delete_book_book_Info)
        print(info_len)
        if info_len>51:
            #4行
            info = ""
            info += self._one_delete_book_book_Info[:17] + "\n"
            info += self._one_delete_book_book_Info[17:34] + "\n"
            info += self._one_delete_book_book_Info[34:51] + "\n"
            info += self._one_delete_book_book_Info[51:68] + "..."
            
        elif 51>=info_len>34:
            #3行
            info = ""
            info += self._one_delete_book_book_Info[:17] + "\n"
            info += self._one_delete_book_book_Info[17:34] + "\n"
            info += self._one_delete_book_book_Info[34:51] + "..."
            
        elif 34>=info_len>17:
            #2行
            info = ""
            info += self._one_delete_book_book_Info[:17] + "\n"
            info += self._one_delete_book_book_Info[17:34] + "..."
            
        else:
            #1行
            info = ""
            info += self._one_delete_book_book_Info[:17] + "..."
            
        self.one_delete_book_bookInfo.set(info)
        info=""
        self.one_delete_book_publicationTime.set(self._one_delete_book_publicationTime)
        self.one_delete_book_isbn.set(self._one_delete_book_isbn)
        self.one_delete_book_inventory.set(str(self._one_delete_book_inventory))
        self.one_delete_book_id.set(self._one_delete_book_id)
        return
    
    def one_delete_book_delete_book_button_click(self):
        if self._one_delete_book_id != '' and self._one_delete_book_author_name != "" and self._one_delete_book_press != '' and self._one_delete_book_book_Info != '' and self._one_delete_book_publicationTime != '' and self._one_delete_book_isbn != '' and self._one_delete_book_inventory != '':
            lb.Del_Book(self._one_delete_book_id)
            messagebox.showinfo("提示", "删除成功")
            self._one_delete_book_author_name = ''
            self._one_delete_book_book_name = ''
            self._one_delete_book_press = ''
            self._one_delete_book_publicationTime = ''
            self._one_delete_book_bookInfo = ''
            self._one_delete_book_isbn = ''
            self._one_delete_book_inventory = ''
            self._one_delete_book_id = ''
            self.goback(self.one_delete_book_window)
            return
        




    def amend_book_info(self):
        self.root.withdraw()
        self.amend_book_window = maliang.Toplevel(self.root,size=(1100,800),title="修改书籍信息")
        self.amend_book_window.center()
        self.amend_book_window.iconbitmap(mypath("favicon.ico"))
        
        self.amend_book_window__Canvas = maliang.Canvas(self.amend_book_window,auto_update=True,expand="xy",keep_ratio="max",auto_zoom=True)
        self.amend_book_window__Canvas.place(x=0,y=0,width=1100,height=800)

        self.amend_book_goback_button = maliang.Button(self.amend_book_window__Canvas,(0,0),size=(50,20),fontsize=15,text="返回", anchor="nw", command=lambda:self.goback(self.amend_book_window))


        def print_option_selected(index):
            print(f"{['书籍搜索',"isbn搜索"][index]} {index}")
        self.amend_book_search_type_OptionButton = maliang.OptionButton(self.amend_book_window__Canvas,(0,50), size=(100,35),fontsize=15,text=("书籍搜索", "ISBN搜索"),command=print_option_selected,default=0)
        self.amend_book_search_button = maliang.Button(self.amend_book_window__Canvas,(100,50),size=(100,35),fontsize=15,text="搜索书籍", anchor="nw", command=lambda:self.amend_book_search_book())
        self.amend_book_search_inputbox = maliang.InputBox(self.amend_book_window__Canvas,(200,50),size=(800,35))
        self.amend_book_search_inputbox.bind("<Return>",self.amend_book_search_book)
        self.amend_book_search_show_tree = ttk.Treeview(self.amend_book_window__Canvas,columns=("书名","作者","出版社","出版时间","ISBN","库存","id"),show="headings")
        self.amend_book_search_show_tree.heading("书名",text="书名")
        self.amend_book_search_show_tree.heading("作者",text="作者")
        self.amend_book_search_show_tree.heading("出版社",text="出版社")
        self.amend_book_search_show_tree.heading("出版时间",text="出版时间")
        self.amend_book_search_show_tree.heading("ISBN",text="ISBN")
        self.amend_book_search_show_tree.heading("库存",text="库存")
        self.amend_book_search_show_tree.heading("id",text="id")
        self.amend_book_search_show_tree.column("书名",width=90)
        self.amend_book_search_show_tree.column("作者",width=90)
        self.amend_book_search_show_tree.column("出版社",width=90)
        self.amend_book_search_show_tree.column("出版时间",width=70)
        self.amend_book_search_show_tree.column("ISBN",width=70)
        self.amend_book_search_show_tree.column("库存",width=50)
        self.amend_book_search_show_tree.column("id",width=90)
        self.amend_book_search_show_tree.place(x=0,y=100,width=1000,height=150)
        self.amend_book_search_show_tree.bind("<ButtonRelease-1>",self.amend_book_search_show_tree_double_click)

        self.amend_book_chang_part_head_text = maliang.Text(self.amend_book_window__Canvas,(0,270),text="修改部分:",underline=True,anchor="nw")

        self.amend_book_chang_part_bookname_text = maliang.Text(self.amend_book_window__Canvas,(0,320),text="书名: ",anchor="nw")
        self.amend_book_chang_part_bookname_inputbox = maliang.InputBox(self.amend_book_window__Canvas,(90,320),size=(830,35))

        self.amend_book_chang_part_author_text = maliang.Text(self.amend_book_window__Canvas,(0,370),text="作者: ",anchor="nw")
        self.amend_book_chang_part_author_inputbox = maliang.InputBox(self.amend_book_window__Canvas,(90,370),size=(830,35))

        self.amend_book_chang_part_press_text = maliang.Text(self.amend_book_window__Canvas,(0,420),text="出版社: ",anchor="nw")
        self.amend_book_chang_part_press_inputbox = maliang.InputBox(self.amend_book_window__Canvas,(120,420),size=(800,35))

        self.amend_book_chang_part_publicationTime_text = maliang.Text(self.amend_book_window__Canvas,(0,470),text="出版时间: ",anchor="nw")
        self.amend_book_chang_part_publicationTime_inputbox = maliang.InputBox(self.amend_book_window__Canvas,(150,470),size=(770,35))

        self.amend_book_chang_part_bookInfo_text = maliang.Text(self.amend_book_window__Canvas,(0,520),text="书籍信息: ",anchor="nw")
        self.amend_book_chang_part_bookInfo_inputbox = tk.Text(self.amend_book_window__Canvas,height=5, width=70)
        self.amend_book_chang_part_bookInfo_inputbox.configure(font=font.Font(family="Helvetica", size=14))
        # self.amend_book_chang_part_bookInfo_inputbox.insert(tk.END, "哈哈哈哈哈哈哈哈哈哈哈啥哈哈哈哈哈哈哈哈哈哈哈哈哈哈哈哈哈哈哈哈哈啥\nhahaha")
        # text_box.get("1.0", tk.END)
        self.amend_book_chang_part_bookInfo_inputbox.place(x=150,y=520)
        
        self.amend_book_chang_part_isbn_text = maliang.Text(self.amend_book_window__Canvas,(0,650),text="ISBN: ",anchor="nw")
        self.amend_book_chang_part_isbn_inputbox = maliang.InputBox(self.amend_book_window__Canvas,(90,650),size=(830,35))

        self.amend_book_chang_part_inventory_text = maliang.Text(self.amend_book_window__Canvas,(0,700),text="库存: ",anchor="nw")
        self.amend_book_chang_part_inventory_inputbox = maliang.InputBox(self.amend_book_window__Canvas,(90,700),size=(830,35))

        self.amend_book_chang_part_id_text = maliang.Text(self.amend_book_window__Canvas,(0,750),text="id: ",anchor="nw")
        self.amend_book_chang_part_id_Text = maliang.Text(self.amend_book_window__Canvas,(90,750))

        self.amend_book_chang_part_bookname_inputbox.bind("<KeyRelease>",self.amend_book_chang_part_input_bookname_change)
        self.amend_book_chang_part_author_inputbox.bind("<KeyRelease>",self.amend_book_chang_part_input_author_change)
        self.amend_book_chang_part_press_inputbox.bind("<KeyRelease>",self.amend_book_chang_part_input_press_change)
        self.amend_book_chang_part_publicationTime_inputbox.bind("<KeyRelease>",self.amend_book_chang_part_input_publicationTime_change)
        self.amend_book_chang_part_bookInfo_inputbox.bind("<KeyRelease>",self.amend_book_chang_part_input_bookInfo_change)
        self.amend_book_chang_part_isbn_inputbox.bind("<KeyRelease>",self.amend_book_chang_part_input_isbn_change)
        self.amend_book_chang_part_inventory_inputbox.bind("<KeyRelease>",self.amend_book_chang_part_input_inventory_change)


        self.amend_book_button = maliang.Button(self.amend_book_window__Canvas,(990,755),size=(100,35),fontsize=15,text="提交修改", anchor="nw", command=lambda:self.amend_book_change_book_info())


        self.amend_book_window.protocol("WM_DELETE_WINDOW", lambda:self.goback(self.amend_book_window))
        pass
    
    def update_init_msg_of_amend_book_msg_to_inputbox(self):
        self.amend_book_chang_part_bookname_inputbox.set(self.amend_book_bookname_first_msg)
        self.amend_book_chang_part_author_inputbox.set(self.amend_book_author_first_msg)
        self.amend_book_chang_part_press_inputbox.set(self.amend_book_press_first_msg)
        self.amend_book_chang_part_publicationTime_inputbox.set(self.amend_book_publicationTime_first_msg)
        self.amend_book_chang_part_bookInfo_inputbox.delete("1.0", tk.END)
        self.amend_book_chang_part_bookInfo_inputbox.insert(tk.END, self.amend_book_bookInfo_first_msg)
        self.amend_book_chang_part_isbn_inputbox.set(self.amend_book_isbn_first_msg)
        self.amend_book_chang_part_inventory_inputbox.set(str(self.amend_book_inventory_first_msg))
        self.amend_book_chang_part_id_Text.set(self.amend_book_id_first_msg)
        pass



    def amend_book_chang_part_input_bookname_change(self,*args):
        if self.amend_book_chang_part_bookname_inputbox.get() != self.amend_book_bookname_first_msg:
            self.amend_book_bookname_second_msg = self.amend_book_chang_part_bookname_inputbox.get()
        elif self.amend_book_chang_part_bookname_inputbox.get() == self.amend_book_bookname_first_msg:
            self.amend_book_bookname_second_msg = ""
        print(self.amend_book_bookname_first_msg+"; "+self.amend_book_bookname_second_msg)
    
    def amend_book_chang_part_input_author_change(self,*args):
        if self.amend_book_chang_part_author_inputbox.get() != self.amend_book_author_first_msg:
            self.amend_book_author_second_msg = self.amend_book_chang_part_author_inputbox.get()
        elif self.amend_book_chang_part_author_inputbox.get() == self.amend_book_author_first_msg:
            self.amend_book_author_second_msg = ""
        print(self.amend_book_author_first_msg+"; "+self.amend_book_author_second_msg)
    
    def amend_book_chang_part_input_press_change(self,*args):
        if self.amend_book_chang_part_press_inputbox.get() != self.amend_book_press_first_msg:
            self.amend_book_press_second_msg = self.amend_book_chang_part_press_inputbox.get()
        elif self.amend_book_chang_part_press_inputbox.get() == self.amend_book_press_first_msg:
            self.amend_book_press_second_msg = ""
        print(self.amend_book_press_first_msg+"; "+self.amend_book_press_second_msg)
    
    def amend_book_chang_part_input_publicationTime_change(self,*args):
        if self.amend_book_chang_part_publicationTime_inputbox.get() != self.amend_book_publicationTime_first_msg:
            self.amend_book_publicationTime_second_msg = self.amend_book_chang_part_publicationTime_inputbox.get()
        elif self.amend_book_chang_part_publicationTime_inputbox.get() == self.amend_book_publicationTime_first_msg:
            self.amend_book_publicationTime_second_msg = ""
        print(self.amend_book_publicationTime_first_msg+"; "+self.amend_book_publicationTime_second_msg)
    
    def amend_book_chang_part_input_bookInfo_change(self,*args):
        if len(self.amend_book_chang_part_bookInfo_inputbox.get("1.0", tk.END).strip()) > 200:
            self.amend_book_chang_part_bookInfo_inputbox.delete("end-2c", tk.END)
        if self.amend_book_chang_part_bookInfo_inputbox.get("1.0", tk.END).strip() != self.amend_book_bookInfo_first_msg:
            self.amend_book_bookInfo_second_msg = self.amend_book_chang_part_bookInfo_inputbox.get("1.0", tk.END).strip()
        elif self.amend_book_chang_part_bookInfo_inputbox.get("1.0", tk.END).strip() == self.amend_book_bookInfo_first_msg:
            self.amend_book_bookInfo_second_msg = ""
        print(self.amend_book_bookInfo_first_msg+"; "+self.amend_book_bookInfo_second_msg)
    
    def amend_book_chang_part_input_isbn_change(self,*args):

        if self.amend_book_chang_part_isbn_inputbox.get() != self.amend_book_isbn_first_msg:
            self.amend_book_isbn_second_msg = self.amend_book_chang_part_isbn_inputbox.get()
        elif self.amend_book_chang_part_isbn_inputbox.get() == self.amend_book_isbn_first_msg:
            self.amend_book_isbn_second_msg = ""
        print(self.amend_book_isbn_first_msg+"; "+self.amend_book_isbn_second_msg)
       
    
    def amend_book_chang_part_input_inventory_change(self,*args):
        true_inventory_geshi = r'^\d+$'
        if bool(re.match(true_inventory_geshi, self.amend_book_chang_part_inventory_inputbox.get())) or self.amend_book_chang_part_inventory_inputbox.get()=="":
            if self.amend_book_chang_part_inventory_inputbox.get() != self.amend_book_inventory_first_msg:
                self.amend_book_inventory_second_msg = self.amend_book_chang_part_inventory_inputbox.get()
            elif self.amend_book_chang_part_inventory_inputbox.get() == self.amend_book_inventory_first_msg:
                self.amend_book_inventory_second_msg = ""
            print("库存修改:"+self.amend_book_inventory_first_msg+"; "+self.amend_book_inventory_second_msg)
        else:
            self.amend_book_chang_part_inventory_inputbox.set(self.amend_book_inventory_first_msg)
    def amend_book_search_book(self,event=None):
        search_thing = self.amend_book_search_inputbox.get()
        print(search_thing)
        search_books_list = []
        if self.amend_book_search_type_OptionButton.get() == 0:
            #书籍搜索
            search_books_list = lb.Find_Books(self.amend_book_search_inputbox.get())
            # print(search_books_list)
            if len(search_books_list)>0:
                self.now_show_amend_books_list = []
                for item in self.amend_book_search_show_tree.get_children():
                    self.amend_book_search_show_tree.delete(item)
                for item in search_books_list:
                    self.now_show_amend_books_list.append(item)
                    self.amend_book_search_show_tree.insert("", "end", values=(item[0],item[1],item[2],item[3],item[5],item[6]))
                    
            else:
                messagebox.showerror("错误", "未找到相关书籍")
            pass

        elif self.amend_book_search_type_OptionButton.get() == 1:
            #isbn搜索
            search_books_dict = lb.Find_book_by_isbn(self.amend_book_search_inputbox.get())
            if search_books_dict['code'] == 200:
                search_books_list = search_books_dict['msg']
                self.now_show_amend_books_list = []
                for item in self.amend_book_search_show_tree.get_children():
                    self.amend_book_search_show_tree.delete(item)
                self.now_show_amend_books_list.append(search_books_list)
                # print(search_books_list)
                self.amend_book_search_show_tree.insert("", "end", values=(search_books_list[0],search_books_list[1],search_books_list[2],search_books_list[3],search_books_list[5],search_books_list[6]))
            elif search_books_dict['code'] == 404:
                messagebox.showerror("错误", "未找到该书籍")
                return 0



    def amend_book_search_show_tree_double_click(self,event=None):
        selected_item = self.amend_book_search_show_tree.selection()[0]
        book_info_in_tree = self.amend_book_search_show_tree.item(selected_item, "values")
        index = self.amend_book_search_show_tree.index(selected_item)
        book_info = self.now_show_amend_books_list[index]
        print(book_info)
        self.amend_book_bookname_first_msg = book_info[0]
        self.amend_book_bookInfo_first_msg = book_info[4]
        self.amend_book_author_first_msg = book_info[1]
        self.amend_book_press_first_msg = book_info[2]
        self.amend_book_publicationTime_first_msg = book_info[3]
        self.amend_book_isbn_first_msg = book_info[5]
        self.amend_book_inventory_first_msg = str(book_info[6])
        self.amend_book_id_first_msg = book_info[7]
        self.update_init_msg_of_amend_book_msg_to_inputbox()

    def amend_book_change_book_info(self):
        update_book_info_list = []
        if self.amend_book_bookname_second_msg != "":
            update_book_info_list.append(self.amend_book_bookname_second_msg)
        else:
            update_book_info_list.append(self.amend_book_bookname_first_msg)
        if self.amend_book_author_second_msg != "":
            update_book_info_list.append(self.amend_book_author_second_msg)
        else:
            update_book_info_list.append(self.amend_book_author_first_msg)
        if self.amend_book_press_second_msg != "":
            update_book_info_list.append(self.amend_book_press_second_msg)
        else:
            update_book_info_list.append(self.amend_book_press_first_msg)
        if self.amend_book_publicationTime_second_msg != "":
            update_book_info_list.append(self.amend_book_publicationTime_second_msg)
        else:
            update_book_info_list.append(self.amend_book_publicationTime_first_msg)
        if self.amend_book_bookInfo_second_msg != "":
            update_book_info_list.append(self.amend_book_bookInfo_second_msg)
        else:
            update_book_info_list.append(self.amend_book_bookInfo_first_msg)
        if self.amend_book_isbn_second_msg != "":
            update_book_info_list.append(self.amend_book_isbn_second_msg)
        else:
            update_book_info_list.append(self.amend_book_isbn_first_msg)
        if self.amend_book_inventory_second_msg == "":
            self.amend_book_inventory_second_msg = 0
        if self.amend_book_inventory_second_msg != "":
            update_book_info_list.append(int(self.amend_book_inventory_second_msg))
        if len(update_book_info_list) == 0:
            messagebox.showerror("错误", "未修改任何信息")
            return 0
        update_book_info_result = lb.amend_book_msg(self.amend_book_id_first_msg,update_book_info_list)
        if update_book_info_result['code'] == 200:
            messagebox.showinfo("提示", "修改成功")
        elif update_book_info_result['code'] == 404:
            messagebox.showerror("错误", update_book_info_result['msg'])
        self.goback(self.amend_book_window)






    def add_student_info(self):
        self.root.withdraw()
        self.add_student_window = maliang.Toplevel(self.root,(1000,600),title="添加学生信息")
        self.add_student_window.center()
        self.add_student_window.iconbitmap(mypath("favicon.ico"))

        self.add_student_window__Canvas = maliang.Canvas(self.add_student_window,auto_update=True,expand="xy",keep_ratio="max",auto_zoom=True)
        self.add_student_window__Canvas.place(x=0,y=0,width=1000,height=600)

        self.add_student_back_button = maliang.Button(self.add_student_window__Canvas,(0,0),size=(50,20),fontsize=15,text="返回", anchor="nw", command=lambda:self.goback(self.add_student_window))

        self.add_student_name_text = maliang.Text(self.add_student_window__Canvas,(0,50),text="姓名：", anchor="nw")
        self.add_student_name_inputbox = maliang.InputBox(self.add_student_window__Canvas,(100,50),size=(500,50),anchor="nw")

        self.add_student_class_text = maliang.Text(self.add_student_window__Canvas,(0,120),text="班级：", anchor="nw")
        self.add_student_class_inputbox = maliang.InputBox(self.add_student_window__Canvas,(100,120),size=(500,50),anchor="nw")
        self.add_student_class_inputbox.bind("<KeyRelease>",self.check_add_student_class_input)

        self.add_student_id_text = maliang.Text(self.add_student_window__Canvas,(0,190),text="学号：", anchor="nw")
        self.add_student_id_inputbox = maliang.InputBox(self.add_student_window__Canvas,(100,190),size=(500,50),anchor="nw")
        self.add_student_id_inputbox.bind("<KeyRelease>",self.check_add_student_id_input)

        self.add_student_qrcode_save_path_text = maliang.Text(self.add_student_window__Canvas,(0,260),text="二维码保存路径：", anchor="nw")
        self.add_student_qrcode_save_path_inputbox = maliang.Text(self.add_student_window__Canvas,(200,260),anchor="nw",text=f"{mypath("学生信息\\")}")
        self.add_student_qrcode_save_path_inputbox.set(f"{mypath('学生信息\\')}")
        
        self.add_student_qrcode_save_path_choise_button = maliang.Button(self.add_student_window__Canvas,(800,260),size=(100,50),text="选择", anchor="nw", command=lambda:self.choose_qrcode_save_path())

        self.add_student_password_text = maliang.Text(self.add_student_window__Canvas,(650,50),text="密码: ", anchor="nw")
        self.add_student_password_inputbox = maliang.InputBox(self.add_student_window__Canvas,(750,50),size=(200,50),anchor="nw",placeholder="(为空随机)")

        def show_qrcode_when_finish_checkbox_change(index):
            print(index)
            self.add_student_show_qrcode_when_finish = index
        self.add_student_show_qrcode_when_finish_checkbox = maliang.CheckBox(self.add_student_window__Canvas,(0,350),anchor="nw", command=show_qrcode_when_finish_checkbox_change)
        self.add_student_show_qrcode_when_finish_checkbox.set(True)
        self.add_student_show_qrcode_when_finish_text = maliang.Text(self.add_student_window__Canvas,(40,350),text="完成后是否显示二维码", anchor="nw")



        self.add_student_add_button = maliang.Button(self.add_student_window__Canvas,(450,530),size=(100,50),text="添加", anchor="nw", command=lambda:self.add_student_info_to_db())






        self.add_student_window.protocol("WM_DELETE_WINDOW", lambda:self.goback(self.add_student_window))
        
        pass

    def check_add_student_class_input(self,*args):
        true_class_geshi = r'^\d+$'
        if not (bool(re.match(true_class_geshi, self.add_student_class_inputbox.get())) or self.add_student_class_inputbox.get()==""):
            self.add_student_class_inputbox.set("")

    def check_add_student_id_input(self,*args):
        true_id_geshi = r'^\d+$'
        if not (bool(re.match(true_id_geshi, self.add_student_id_inputbox.get())) or self.add_student_id_inputbox.get()==""):
            self.add_student_id_inputbox.set("")

    def choose_qrcode_save_path(self):
        path = filedialog.askdirectory()
        if path:
            self.add_student_qrcode_save_path_inputbox.set(path)
            self.add_student_qrcode_save_path = path
        
        del path

    def add_student_info_to_db(self):
        student_name = self.add_student_name_inputbox.get()
        student_class = self.add_student_class_inputbox.get()
        student_id = self.add_student_id_inputbox.get()
        student_password = self.add_student_password_inputbox.get()
        student_qr_save_path = self.add_student_qrcode_save_path_inputbox.get()
        if student_qr_save_path == "":
            student_qr_save_path = mypath('学生信息\\')
        if not(student_name and not student_name.isspace() and student_class and not student_class.isspace() and student_id and not student_id.isspace()):
            messagebox.showerror("错误", "姓名、班级或学号不能为空")
            return 0
        student_name = student_name.strip()
        student_class = student_class.strip()
        student_id = student_id.strip()
        student_password = student_password.strip()
        if os.path.exists(student_qr_save_path):
            print(f"路径 {student_qr_save_path} 已存在。")
        else:
            try:
                # 尝试创建路径
                os.makedirs(student_qr_save_path)
                print(f"路径 {student_qr_save_path} 不存在，已创建。")
            except OSError as e:
                print(f"创建路径 {student_qr_save_path} 时出错: {e}")

        if lb.Register_User(student_name, student_class, student_id, student_password,self.add_student_show_qrcode_when_finish,student_qr_save_path):
            messagebox.showinfo("提示", "添加成功")
        else:
            messagebox.showerror("错误", "添加失败")
        self.goback(self.add_student_window)
        return 0



    def delete_student_info(self):
        self.root.withdraw()
        self.delete_student_window = maliang.Toplevel(self.root,(1000,800),title="注销学生信息")
        self.delete_student_window.center()
        self.delete_student_window.iconbitmap(mypath("favicon.ico"))

        self.delete_student_window__Canvas = maliang.Canvas(self.delete_student_window,auto_update=True,expand="xy",keep_ratio="max",auto_zoom=True)
        self.delete_student_window__Canvas.place(x=0,y=0,width=1000,height=800)

        self.delete_student_back_button = maliang.Button(self.delete_student_window__Canvas,(0,0),size=(50,20),fontsize=15,text="返回", anchor="nw", command=lambda:self.goback(self.delete_student_window))

        def print_option_selected(index):
            print(f"{['用户名搜索',"班级搜索","座号搜索","二维码扫描","扫描枪扫描"][index]} {index}")
        self.delete_student_search_type_OptionButton = maliang.OptionButton(self.delete_student_window__Canvas,(0,50), size=(100,35),fontsize=15,text=("用户名搜索", "班级搜索", "座号搜索","二维码扫描","扫描枪扫描"),command=print_option_selected,default=0)
        self.delete_student_search_button = maliang.Button(self.delete_student_window__Canvas,(100,50),size=(100,35),fontsize=15,text="搜索", anchor="nw", command=lambda:self.delete_student_search_student())
        self.delete_student_inputbox = maliang.InputBox(self.delete_student_window__Canvas,(200,50),size=(800,35))
        self.delete_student_inputbox.bind("<Return>",self.delete_student_search_student)
        self.delete_student_opencv_button = maliang.Button(self.delete_student_window__Canvas,(0,100),size=(200,50),fontsize=18,text="扫描学生二维码", anchor="nw", command=lambda:self.delete_student_opencv_for_student())

        self.delete_student_show_students_tree = ttk.Treeview(self.delete_student_window__Canvas,columns=("姓名","班级","座号"),show="headings")
        self.delete_student_show_students_tree.heading("姓名", text="姓名")
        self.delete_student_show_students_tree.heading("班级", text="班级")
        self.delete_student_show_students_tree.heading("座号", text="座号")
        self.delete_student_show_students_tree.column("姓名", width=100)
        self.delete_student_show_students_tree.column("班级", width=100)
        self.delete_student_show_students_tree.column("座号", width=100)
        self.delete_student_show_students_tree.place(x=210,y=100,width=750,height=500)
        self.delete_student_show_students_tree.bind("<ButtonRelease-1>",self.delete_student_show_students_tree_click)

        self.delete_student_student_msg_part_student_name_text = maliang.Text(self.delete_student_window__Canvas,(0,170),text="姓名：", anchor="nw")
        self.delete_student_student_msg_part_student_name_Text = maliang.Text(self.delete_student_window__Canvas,(50,170),anchor="nw")

        self.delete_student_student_msg_part_student_class_text = maliang.Text(self.delete_student_window__Canvas,(0,200),text="班级：", anchor="nw")
        self.delete_student_student_msg_part_student_class_Text = maliang.Text(self.delete_student_window__Canvas,(50,200),anchor="nw")

        self.delete_student_student_msg_part_student_id_text = maliang.Text(self.delete_student_window__Canvas,(0,230),text="座号：", anchor="nw")
        self.delete_student_student_msg_part_student_id_Text = maliang.Text(self.delete_student_window__Canvas,(50,230),anchor="nw")

        self.delete_student_student_msg_part_student_borrowbooks_text = maliang.Text(self.delete_student_window__Canvas,(0,260),text="正在借阅的图书：", anchor="nw")
        self.delete_student_student_msg_part_student_borrowbooks_tree = ttk.Treeview(self.delete_student_window__Canvas,columns=("书名"),show="headings")
        self.delete_student_student_msg_part_student_borrowbooks_tree.heading("书名", text="书名")
        self.delete_student_student_msg_part_student_borrowbooks_tree.column("书名", width=200)
        self.delete_student_student_msg_part_student_borrowbooks_tree.place(x=5,y=290,width=200,height=150)

        self.delete_student_student_msg_part_student_borrowedbooks_text = maliang.Text(self.delete_student_window__Canvas,(0,450),text="已借阅的图书：", anchor="nw")
        self.delete_student_student_msg_part_student_borrowedbooks_tree = ttk.Treeview(self.delete_student_window__Canvas,columns=("书名"),show="headings")
        self.delete_student_student_msg_part_student_borrowedbooks_tree.heading("书名", text="书名")
        self.delete_student_student_msg_part_student_borrowedbooks_tree.column("书名", width=200)
        self.delete_student_student_msg_part_student_borrowedbooks_tree.place(x=5,y=480,width=200,height=310)

        self.delete_student_delete_button = maliang.Button(self.delete_student_window__Canvas,(500,750),size=(100,50),fontsize=15,text="注销", anchor="nw", command=lambda:self.delete_student_button_click())
        
        self.delete_student_window.protocol("WM_DELETE_WINDOW", lambda:self.goback(self.delete_student_window))
    

    def delete_student_search_student(self,event=None):
        print(self.delete_student_inputbox.get())
        if self.delete_student_search_type_OptionButton.get() == 0:
            result = lb.find_user_by_name(self.delete_student_inputbox.get())
            for item in self.delete_student_show_students_tree.get_children():
                self.delete_student_show_students_tree.delete(item)
            self.delete_student_now_show_students_list = []
            for item in result:
                self.delete_student_show_students_tree.insert("", "end", values=(item[0],item[2],item[1]))
                self.delete_student_now_show_students_list.append(item)
            print(result)
            pass
        elif self.delete_student_search_type_OptionButton.get() == 1:
            result = lb.find_user_by_class(self.delete_student_inputbox.get())
            for item in self.delete_student_show_students_tree.get_children():
                self.delete_student_show_students_tree.delete(item)
            self.delete_student_now_show_students_list = []
            for item in result:
                self.delete_student_show_students_tree.insert("", "end", values=(item[0],item[2],item[1]))
                self.delete_student_now_show_students_list.append(item)
            print(result)
            pass
        elif self.delete_student_search_type_OptionButton.get() == 2:
            result = lb.find_user_by_id(self.delete_student_inputbox.get())
            for item in self.delete_student_show_students_tree.get_children():
                self.delete_student_show_students_tree.delete(item)
            self.delete_student_now_show_students_list = []
            for item in result:
                self.delete_student_show_students_tree.insert("", "end", values=(item[0],item[2],item[1]))
                self.delete_student_now_show_students_list.append(item)
            print(result)
            pass
            
        elif self.delete_student_search_type_OptionButton.get() == 4:
            student_msg = self.delete_student_inputbox.get()
            jianli_lianjie()
            student_msg = lb.Decrypt_User_Info(student_msg)
            student_msg = lb.Login_User(student_msg[0], student_msg[2], student_msg[1], student_msg[3])
            print(student_msg)
            if student_msg['code'] == 200:
                for item in self.delete_student_show_students_tree.get_children():
                    self.delete_student_show_students_tree.delete(item)
                self.delete_student_now_show_students_list = []
                student_msg = student_msg['msg'][0]
                print(student_msg)
                name = student_msg[0]
                class_ = student_msg[2]
                id_ = student_msg[1]
                borrowbooks = student_msg[3]
                borrowedbooks = student_msg[5]
                password = student_msg[4]
                self.delete_student_show_students_tree.insert("", "end", values=(name,class_,id_))
                self.delete_student_now_show_students_list.append(student_msg)
            else:
                for item in self.delete_student_show_students_tree.get_children():
                    self.delete_student_show_students_tree.delete(item)
                self.delete_student_now_show_students_list = []
                messagebox.showerror("错误", "未找到该学生")



    def delete_student_opencv_for_student(self):
        if self.delete_student_search_type_OptionButton.get() == 3:
            student_msg = lb.cv_for_student()
            jianli_lianjie()
            student_msg = lb.Login_User(student_msg[0], student_msg[2], student_msg[1], student_msg[3])
            print(student_msg)
            if student_msg['code'] == 200:
                student_msg = student_msg['msg'][0]
                print(student_msg)
                name = student_msg[0]
                class_ = student_msg[2]
                id_ = student_msg[1]
                borrowbooks = student_msg[3]
                borrowedbooks = student_msg[5]
                password = student_msg[4]
                self.delete_student_show_students_tree.insert("", "end", values=(name,class_,id_))
                self.delete_student_now_show_students_list.append(student_msg)
                
        else:
            for item in self.delete_student_show_students_tree.get_children():
                self.delete_student_show_students_tree.delete(item)
            self.delete_student_now_show_students_list = []
            messagebox.showerror("错误", "未找到该学生")


    def delete_student_show_students_tree_click(self,event):
        selected_item = self.delete_student_show_students_tree.selection()[0]
        book_info_in_tree = self.delete_student_show_students_tree.item(selected_item, "values")
        index = self.delete_student_show_students_tree.index(selected_item)
        book_info = self.delete_student_now_show_students_list[index]
        print(book_info)
        self.delete_student_name = book_info[0]
        self.delete_student_class = book_info[2]
        self.delete_student_id = book_info[1]
        self.delete_student_borrowbooks = json.loads(book_info[3])
        self.delete_student_borrowedbooks = json.loads(book_info[5])
        self.delete_student_password = book_info[4]
        self.update_delete_student_msg_part()

    def update_delete_student_msg_part(self):
        for item in self.delete_student_student_msg_part_student_borrowbooks_tree.get_children():
            self.delete_student_student_msg_part_student_borrowbooks_tree.delete(item)
        for item in self.delete_student_student_msg_part_student_borrowedbooks_tree.get_children():
            self.delete_student_student_msg_part_student_borrowedbooks_tree.delete(item)
        self.delete_student_student_msg_part_student_name_Text.set(str(self.delete_student_name))
        self.delete_student_student_msg_part_student_class_Text.set(str(self.delete_student_class))
        self.delete_student_student_msg_part_student_id_Text.set(str(self.delete_student_id))
        for borrowbook in self.delete_student_borrowbooks:
            msg = lb.Find_book_by_isbn(borrowbook)
            if msg['code'] == 200:
                msg = msg['msg'][0]
                self.delete_student_student_msg_part_student_borrowbooks_tree.insert("", "end", values=(msg))
            else:
                self.delete_student_student_msg_part_student_borrowbooks_tree.insert("", "end", values=(f"{borrowbook} (搜索失败)"))
        for borrowedbook in self.delete_student_borrowedbooks:
            msg = lb.Find_book_by_isbn(borrowedbook)
            if msg['code'] == 200:
                msg = msg['msg'][0]
                self.delete_student_student_msg_part_student_borrowedbooks_tree.insert("", "end", values=(msg))
            else:
                self.delete_student_student_msg_part_student_borrowedbooks_tree.insert("", "end", values=(f"{borrowedbook} (搜索失败)"))
    
    def delete_student_button_click(self):
        if messagebox.askyesno("提示", "确认注销该学生信息？"):
            name =self.delete_student_name
            id_ = self.delete_student_id
            class_ = self.delete_student_class
            password = self.delete_student_password
            lb.Delete_User(name, id_, class_, password)
            messagebox.showinfo("提示", "注销成功")
            self.goback(self.delete_student_window)
    def amend_student_info(self):
        self.root.withdraw()
        self.amend_student_window = maliang.Toplevel(self.root,(1000,800),title="修改学生信息")
        self.amend_student_window.iconbitmap(mypath("favicon.ico"))
        self.amend_student_window.center()

        













        self.amend_student_window.protocol("WM_DELETE_WINDOW", lambda:self.goback(self.amend_student_window))  






        pass

    def borrow_and_return_query(self):
        pass
    def export_log(self):
        pass
Developer = LibrarySystem()

