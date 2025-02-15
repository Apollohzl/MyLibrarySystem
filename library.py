#导入所需库
import tkinter as tk
from tkinter import messagebox
import json
from datetime import datetime
from tkinter import filedialog
import os
import openpyxl
from PIL import Image
import pyzbar.pyzbar as pyzbar
import barcode
from barcode.writer import ImageWriter
import hashlib
import base64
import time
import datetime
import sqlite3
from encrypt import encrypt
import random
import cv2
import threading
import queue
def mypath(other: str | None = ""):
    return str(os.path.dirname(os.path.abspath(__file__)))+"\\"+other
#登录密码加密
def mimajiami(text, key="Apollokey"):
    # 将key与text拼接，确保算法可逆
    text = key + text
    # SHA-256哈希
    hash_object = hashlib.sha256(text.encode())
    hash_hex = hash_object.hexdigest()
    
    # Base64编码
    encoded = base64.b64encode(hash_hex.encode()).decode()
    
    # 调整长度到固定长度，例如18字符
    fixed_length = 18
    if len(encoded) > fixed_length:
        encoded = encoded[:fixed_length]
    else:
        encoded = encoded.ljust(fixed_length, '0')  # 使用0填充到18字符
    
    return encoded

global Librarysql
global systemlog
global cursor1
global cursor2
def jianli_lianjie(a=False):
    global Librarysql
    global systemlog
    global cursor1
    global cursor2
    #建立数据库
    Librarysql = sqlite3.connect(mypath("Library.oflibrary"))
    systemlog = sqlite3.connect(mypath("Library.log"))
    if a==True:
        cursor1 = Librarysql.cursor()
        cursor2 = systemlog.cursor()
    #建立表
    try:
        #图书馆书籍库
        tosql = '''
        Create table books(
            bookname text,
            author text,
            press text,
            publicationTime text,
            bookInfo text,
            isbn text,
            inventory int,
            id text);
        '''
        Librarysql.execute(tosql)
        #现已借出的书籍(会修改信息，还书时会删除信息)
        tosql = '''
        Create table borrow(
            bookname text,
            author text,
            press text,
            publicationTime text,
            isbn text,
            borrowtime text,
            MustReturnTime text,
            ReaderName text,
            Readerid int,
            Readerclass int
            );
        '''
        Librarysql.execute(tosql)
        #BorrowedBooks籍(不会修改信息，只要有借过书的都有记录)
        tosql = '''
        Create table borrowhistory(
            bookname text,
            author text,
            press text,
            publicationTime text,
            isbn text,
            borrowtime text,
            MustBookReturnTime text,
            TrueBookReturnTime text,
            ReaderName text,
            Readerid int,
            Readerclass int
            );
        '''
        Librarysql.execute(tosql)
        tosql = '''
        Create table users(
            Username text,
            Userid int,
            Userclass int,
            UserBorrowBooks text,
            UserPassword text,
            UserBorrowedBooks text
            );
        '''
        Librarysql.execute(tosql)
    except Exception as error:
        print(error)

    try:
        tosql = '''
        Create table log(
            Time text,
            Do text
            );
        '''
        systemlog.execute(tosql)
    except Exception as error:
        print(error)
jianli_lianjie()
class Library:
    def __init__(self):
        self.info_queue = queue.Queue()
        pass
#日志添加
    def Add_Log(self,log):
        print(f">Library>Add_Log({log})")
        tosql = '''
        INSERT into log values(?,?)
        '''
        msgtosql = (
            datetime.datetime.now(),
            log)
        systemlog.execute(tosql,msgtosql)
        systemlog.commit()
        print(f">Library>Add_Log()-R:Log Add Ok!")

#借书历史添加
    def Add_Borrow_History(self,book,isbn):
        print(f">Library>Add_Borrow_History({book})")
        tosql = '''INSERT into borrowhistory values(?,?,?,?,?,?,?,?,?,?,?)'''
        msgtosql = (
            book[0],
            book[1],
            book[2],
            book[3],
            isbn,
            book[5],
            book[6],
            "",
            book[7],
            book[8],
            book[9])
        Librarysql.execute(tosql,msgtosql)
        Librarysql.commit()
        print(f">Library>Add_Borrow_History()-R:Borrow History Add Ok!")
#还书历史修改
    def Add_Return_History(self,isbn):
        print(f">Library>Add_Return_History({isbn}->type:{type(isbn)})")
        tosql = '''
        UPDATE borrowhistory set TrueBookReturnTime =? where isbn =?
        '''
        msgtosql = (
            datetime.datetime.now(),
            isbn)
        Librarysql.execute(tosql,msgtosql)
        print(Librarysql.commit())
        print(f">Library>Add_Return_History()-R:Return History Add Ok!")
#User信息卡添加
    def Add_User_Readsbookslist(self,user,readingbooks:list,borrowedbooks:list):
        print(f">Library>Add_User_Readsbookslist({user},{readingbooks},{borrowedbooks})")
        tosql = '''
        UPDATE users set UserBorrowBooks=? where Username =? and Userid =? and Userclass =? and UserPassword =?
        '''
        msgtosql = (
            json.dumps(readingbooks),
            user[0],
            user[1],
            user[2],
            user[4]
        )
        Librarysql.execute(tosql,msgtosql)
        Librarysql.commit()
        tosql = '''
        UPDATE users set UserBorrowedBooks=? where Username =? and Userid =? and Userclass =? and UserPassword =?
        '''
        msgtosql = (
            json.dumps(borrowedbooks),
            user[0],
            user[1],
            user[2],
            user[4]
        )
        Librarysql.execute(tosql,msgtosql)
        Librarysql.commit()


#==============================================================================
#书籍类组件

#添加书籍
    def Add_Book(self, book:list):
        print(f">Library>Add_Book({book})->Library>Save_Book(book)")
        self.Add_Log(f"添加书籍 {book} 成功")
        self.Save_Book(book)
    
    def Save_Book(self,book:list):
        print(">Library>Save_Book()")
        tosql = '''
        INSERT into books values(?,?,?,?,?,?,?,?)
        '''
        ti = ""
        for i in str(time.time()).split('.'):
            ti += i
        msgtosql = (
            book[0],
            book[1],
            book[2],
            book[3],
            book[4],
            book[5],
            book[6],
            ti)
        Librarysql.execute(tosql,msgtosql)
        Librarysql.commit()
        print(">Library>Save_Book()-R:Save Ok!")
#excel导入书籍
    def Import_Book_From_Excel(self,path):
            print(f">Library>Import_Book_From_Excel({path})")
        
            workbook = openpyxl.load_workbook(path)
            ws = workbook.active
            for i in range(2,ws.max_row+1):
                A = str(ws.cell(column=2,row=i).value)
                B = str(ws.cell(column=4,row=i).value)
                C = str(ws.cell(column=6,row=i).value)
                D = str(ws.cell(column=7,row=i).value)
                E = str(ws.cell(column=11,row=i).value)
                F = str(ws.cell(column=12,row=i).value)
                G = str(ws.cell(column=5,row=i).value)
                F = ws.cell(column=13,row=i).value
                
                if A!='None' and B!='None' and C!='None' and D!='None' and E!='None' and F!='None' and G!='None':
                    book = [
                        A,
                        B,
                        C,
                        D,
                        E,
                        G,
                        F
        
                    ]
                    self.Add_Book(book)
                    time.sleep(0.01)
                    self.Add_Log(f"导入书籍 {path}")
            print(f">Library>Import_Book_From_Excel()-R:Import Ok!")
        # except Exception as error:
        #     print(f">Library>Import_Book_From_Excel()-R:Import Error:{error}")


#删除书籍
    def Delete_Book(self,id):
        print(f">Library>Del_Book({id})->Library>Del_Book(book)")
        self.Add_Log(f"删除书籍 {id}")
        self.Del_Book(id)

    def Del_Book(self,id):
        print(">Library>Del_Book()")
        tosql = '''
        DELETE from books where id =?
        '''
        Librarysql.execute(tosql,(id,))
        Librarysql.commit()
        self.Add_Log(f"删除书籍 {id}")
        print(">Library>Del_Book()-R:Del Ok!")

#列出书籍
    def List_Book(self):
        print(">Library>List_Book()")
        tosql = '''
        SELECT * from books
        '''
        cursor = Librarysql.execute(tosql)
        return cursor.fetchall()
    
#列出借出过的书
    def List_Borrowed_Book(self)->list:
        print(">Library>List_Borrowed_Book()")
        tosql = '''
        SELECT * from borrowhistory
        '''
        cursor = Librarysql.execute(tosql)
        return cursor.fetchall()
#判断用户是否借过这本书了
    def Login_User_Has_Book(self,usermsg:list,isbn:str)->bool:
        print(f">Library>Login_User_Has_Book({usermsg},{isbn})")
        tosql = '''
        SELECT * from users where Username =? and Userid =? and Userclass =? and UserPassword =?
        '''
        cursor = Librarysql.execute(tosql,(usermsg[0],usermsg[1],usermsg[2],usermsg[4]))
        user = cursor.fetchall()
        if len(user) == 1:
            user = user[0]
            if user[3] == "[]":
                return False
            else:
                userborrowedbooks = json.loads(user[3])
                if isbn in userborrowedbooks:
                    return True
                else:
                    return False
        else:
            return False
#借书
    def Borrow_Book(self,isbn:str,usermsg:list,save_history=False):
        print(f">Library>Borrow_Book({isbn})")
        #检查inventory并更新inventory
        tosql = '''
        SELECT * from books where isbn =?
        '''
        
        cursor = Librarysql.execute(tosql,(isbn,))
        BookMsg = cursor.fetchall()
        if len(BookMsg) == 1:
            BookMsg = BookMsg[0]
            if BookMsg[6] == 0:
                print(f">Library>Borrow_Book()-R:Have No More {isbn} Book!")
                return "这本书借完了"
            else:
                if usermsg != []:
                    if self.Login_User_Has_Book(usermsg,isbn) == False:
                        #修改inventory
                        tosql = '''
                        UPDATE books SET inventory =? where isbn =?
                        '''
                        Librarysql.execute(tosql,(BookMsg[6]-1,isbn))
                        Librarysql.commit()

                        #添加借书borrow历史记录
                        if save_history:
                            tosql = '''
                            INSERT into borrow values(?,?,?,?,?,?,?,?,?,?)
                            '''
                            msgtosql = (
                                BookMsg[0],
                                BookMsg[1],
                                BookMsg[2],
                                BookMsg[3],
                                isbn,
                                datetime.datetime.now(),
                                datetime.datetime.now()+datetime.timedelta(seconds=20),
                                usermsg[0],
                                usermsg[1],
                                usermsg[2])
                            Librarysql.execute(tosql,msgtosql)
                            Librarysql.commit()

                            #添加借书历史borrowhistory记录
                            self.Add_Borrow_History(msgtosql,isbn)
                            #添加User借阅信息
                            newreadbookslist = json.loads(usermsg[3])
                            borrowedbookslist = json.loads(usermsg[5])
                            print("===newreadbookslist========")
                            print(newreadbookslist)
                            print(type(newreadbookslist))
                            print("===newreadbookslist========")
                            newreadbookslist.append(isbn)
                            if isbn not in borrowedbookslist:
                                borrowedbookslist.append(isbn)
                            print("=add==newreadbookslist========")
                            print(newreadbookslist)
                            print("================================")
                            self.Add_User_Readsbookslist(usermsg,newreadbookslist,borrowedbookslist)
                            #添加User借阅记录
                            self.Add_Log(f"User {usermsg} 借阅了 {BookMsg} 书籍")
                        print(f">Library>Borrow_Book()-R:Borrow {isbn} Book!")
                        return "借书成功"
                    else:
                        print(f">Library>Borrow_Book()-R:You Have Already Borrowed {isbn} Book!")
                        return "你已经借过这本书了"
                elif usermsg == []:
                    #修改inventory
                    tosql = '''
                    UPDATE books SET inventory =? where isbn =?
                    '''
                    Librarysql.execute(tosql,(BookMsg[6]-1,isbn))
                    Librarysql.commit()
                    if save_history:
                        tosql = '''
                            INSERT into borrow values(?,?,?,?,?,?,?,?,?,?)
                            '''
                        msgtosql = (
                            BookMsg[0],
                            BookMsg[1],
                            BookMsg[2],
                            BookMsg[3],
                            isbn,
                            datetime.datetime.now(),
                            datetime.datetime.now()+datetime.timedelta(seconds=20),
                            "老师",
                            "",
                            "")
                        Librarysql.execute(tosql,msgtosql)
                        Librarysql.commit()
                        self.Add_Log(f"User 老师 借阅了 {BookMsg} 书籍")
                        #添加借书历史borrowhistory记录
                        self.Add_Borrow_History(msgtosql,isbn)

                    print(f">Library>Borrow_Book()-R:Borrow {isbn} Book!")
                    return "借书成功"
        else:
            print(f">Library>Borrow_Book()-R:No {isbn} Book!")
            return "没有这本书"
        

#还书
    def Return_Book(self,isbn,usermsg:list,save_history=False):
        print(f">Library>Return_Book({isbn})")
        tosql = '''
        SELECT * from books where isbn =?
        '''
        cursor = Librarysql.execute(tosql,(isbn,))
        BookMsg = cursor.fetchall()
        print(f"BookMsg={BookMsg}")
        if len(BookMsg)!= 0:
            if usermsg != []:
                #修改inventory
                tosql = '''
                UPDATE books SET inventory =? where isbn =?
                '''
                Librarysql.execute(tosql,(BookMsg[0][6]+1,isbn))
                Librarysql.commit()
                if save_history:
                    print("删除还书记录")
                    #删除还书记录borrow
                    tosql = '''DELETE from borrow where isbn =? and Readername=? and Readerid=? and Readerclass=?'''
                    msgtosql = (isbn,usermsg[0],usermsg[1],usermsg[2])
                    Librarysql.execute(tosql,msgtosql)
                    Librarysql.commit()
                    print("========================")
                    #添加User还书记录
                    self.Add_Return_History(isbn)
                    newreadbookslist = json.loads(usermsg[3])
                    newreadbookslist.remove(isbn)
                    borrowedbookslist = json.loads(usermsg[5])
                    self.Add_Log(f"User {usermsg} 还了 {BookMsg} 书籍")
                self.Add_User_Readsbookslist(usermsg,newreadbookslist,borrowedbookslist)
                return "还书成功"
            else:
                #修改inventory
                tosql = '''
                UPDATE books SET inventory =? where isbn =?
                '''
                Librarysql.execute(tosql,(BookMsg[0][6]+1,isbn))
                Librarysql.commit()
                if save_history:
                    print("删除还书记录")
                    #删除还书记录borrow
                    tosql = '''DELETE from borrow where isbn =? and Readername=? and Readerid=? and Readerclass=?'''
                    msgtosql = (isbn,"老师","","")
                    Librarysql.execute(tosql,msgtosql)
                    Librarysql.commit()
                    print("========================")
                return "还书成功"
        else:
            print(f">Library>Return_Book()-R:No {isbn} Book!")
            return "没有这本书"

#搜索书籍
    def Find_Books(self,searchname)->list[tuple] :
        print(f">Library>Find_Books({searchname})")
        results = []
        tosql = '''SELECT * from books'''
        cursor= Librarysql.execute(tosql)
        books = cursor.fetchall()
        if searchname != "":
            #一级搜索：全搜索
            for book in books:
                if (searchname in book[0]) or (searchname in book[1]) or (searchname in book[2]) or (searchname in book[3]):
                    results.append(book)
            #二级搜索：二字
            for book in books:
                    for i in range(len(searchname)-1):
                        ToSname = searchname[i]+searchname[i+1]
                        if (ToSname in book[0]) or (ToSname in book[1]) or (ToSname in book[2]) or (searchname in book[3]):
                            if book not in results:
                                results.append(book)
            #三级搜索：挨个字
            for book in books:
                for word in searchname:
                    if (word in book[0]) or (word in book[1]) or (word in book[2]) or (searchname in book[3]):
                        if book not in results:
                            results.append(book)
            self.Add_Log(f"User搜索 {searchname}")
            return results
        else:
            return books
    
    def Find_book_by_isbn(self,isbn:int)->dict:
        print(f">Library>Find_book_by_isbn({isbn})")
        tosql = '''SELECT * from books where isbn =?'''
        cursor= Librarysql.execute(tosql,(isbn,))
        book = cursor.fetchall()
        if len(book)!= 0:
            return {"code":200,"msg":book[0]}
        else:
            return {"code":404,"msg":"扫描不清楚或无该图书，请核对isbn码是否破损"}



    def amend_book_msg(self,book_id,book_new:list):
        print(f">Library>amend_book_msg({book_id},{book_new})")
        try:
            tosql = '''
            UPDATE books set bookname =?,author =?,press =?,publicationTime =?,bookInfo =?,isbn =?,inventory =? where id =?
            '''
            msgtosql = (
                book_new[0],
                book_new[1],
                book_new[2],
                book_new[3],
                book_new[4],
                book_new[5],
                book_new[6],
                book_id)
            Librarysql.execute(tosql,msgtosql)
            Librarysql.commit()
            self.Add_Log(f"User 修改书籍{book_id}信息 ")
            return {'code':200,'msg':"修改成功"}
        except Exception as e:
            print(f">Library>amend_book_msg()-R:Error:{e}")
            return {'code':404,'msg':f"修改失败 {e}"}
#==============================================================================
#User类组件


#User的注册
    def Register_User(self,username,userclass,userid,password,open=False,save_path=f"{mypath('学生信息\\')}"):
        print(f">Library>Register_User({username},{userid},{userclass},{password},openimg?={open},password自定义?={password},save_path={save_path})")
        if password == True or password == "":
            firstpassword = encrypt.一级加密(str(random.randint(100000,999999)))
        elif type(password)==type("123456"):
            firstpassword = mimajiami(password)
        else:
            print(">Library>Register_User()-R:password参数 error!")
            return False
        tosql = '''
        INSERT into users values(?,?,?,?,?,?)
        '''
        msgtosql = (
            username,
            userid,
            userclass,
            "[]",
            firstpassword,
            "[]")
        Librarysql.execute(tosql,msgtosql)
        Librarysql.commit()
        self.Add_Log(f"User {username} 注册")
        if save_path== "":
            encrypt.自动化加密并二维码(username,userclass,userid,firstpassword,open)
        else:
            encrypt.自动化加密并二维码(username,userclass,userid,firstpassword,open,save_path)
        print(f">Library>Register_User()-R:Register Ok!")
        return True
#User信息解密-<str
    def Decrypt_User_Info(self,text:str)->list:
        User_msg = encrypt.解密(text)
        return User_msg
#User的验证
    def Login_User(self, username: str, userid: int, userclass: int, password: str) -> list:
        print(f">Library>Login_User(name={username},class={userclass},id={userid},p={password})")
        # 加密用户输入的密码
        encrypted_password = password
        username = str(username)
        userclass = int(userclass)
        userid = int(userid)
        print(f">Library>Login_User({type(username)},{type(userclass)},{type(userid)},{type(password)})")
        tosql = '''
        SELECT * from users where Username =? and Userclass =? and Userid =? and UserPassword =?
        '''
        msgtosql = (
            username,
            userclass,
            userid,
            encrypted_password)  # 使用加密后的密码进行查询
        print(msgtosql)
        cursor = Librarysql.execute(tosql, msgtosql)
        user = cursor.fetchall()
        """
        Username text,
        Userid int,
        Userclass int,
        UserBorrowedBooks text,
        UserPassword text"""
        print(user)

        # 添加调试信息
        if len(user) == 0:
            print("No matching user found. Checking individual fields...")
            cursor.execute("SELECT * FROM users WHERE Username = ?", (username,))
            user_by_username = cursor.fetchall()
            print(f"Users with matching username: {user_by_username}")
            cursor.execute("SELECT * FROM users WHERE Userclass = ?", (userclass,))
            user_by_class = cursor.fetchall()
            print(f"Users with matching userclass: {user_by_class}")
            cursor.execute("SELECT * FROM users WHERE Userid = ?", (userid,))
            user_by_id = cursor.fetchall()
            print(f"Users with matching userid: {user_by_id}")
            for msg in self.List_User():
                print(msg)
            print(f"password={encrypted_password}")
            cursor.execute("SELECT * FROM users WHERE UserPassword = ?", (encrypted_password,))
            user_by_password = cursor.fetchall()
            print(f"Users with matching password: {user_by_password}")

        if len(user) == 1:
            self.Add_Log(f"User {username} 验证通过")
            return {"code": 200, "msg": user}
        else:
            return {"code": 404, "msg": "密码错误或User名错误,或二维码错误,或User不存在,请联系老师"}

#UserPassword修改
    def Change_Password(self,username,oldpassword,newpassword):
        print(f">Library>Change_Password({username},{oldpassword},{newpassword})")
        oldpassword = mimajiami(oldpassword)
        newpassword = mimajiami(newpassword)
        tosql = '''
        UPDATE users set UserPassword =? where Username =? and UserPassword =?'''
        msgtosql = (
            newpassword,
            username,
            oldpassword)
        Librarysql.execute(tosql,msgtosql)
        Librarysql.commit()
        self.Add_Log(f"User {username} 修改密码")

#User列表
    def List_User(self):
        print(">Library>List_User()")
        tosql = '''
        SELECT * from users
        '''
        cursor = Librarysql.execute(tosql)
        return cursor.fetchall()

#User搜索name
    def find_user_by_name(self,name)->list:
        print(f">Library>find_user_by_name({name})")
        tosql = '''
        SELECT * from users`
        '''
        cursor = Librarysql.execute(tosql)
        users = cursor.fetchall()
        print(users)


#注销User，提供User名，id，class注销
    def Delete_User(self,username,userid,userclass):
        print(f">Library>Delete_User({username},{userid},{userclass})")
        tosql = '''
        DELETE from users where Username =? and Userid =? and Userclass =?
        '''
        msgtosql = (
            username,
            userid,
            userclass)
        Librarysql.execute(tosql,msgtosql)
        Librarysql.commit()
        self.Add_Log(f"User {username} 注销")
        print(f">Library>Delete_User()-R:Delete Ok!")

#list正在借书
    def List_Borrowing(self)->list:
        print(">Library>List_Borrowing()")
        tosql = '''
        SELECT * from borrow
        '''
        cursor = Librarysql.execute(tosql)
        return cursor.fetchall()
    

    def cv_for_student(self)->list:
        global thread 
        thread = threading.Thread(target=self._cv_for_student)
        thread.start()
        thread.join()
        return self.info_queue.get()
    
    def _cv_for_student(self):
        global cursor1
        global thread
        print("打开摄像头扫描学生码")
        cap = cv2.VideoCapture(0)

        if not cap.isOpened():
            print("Error: Could not open video source.")
            self.info_queue.put([])
            return
        jianli_lianjie(True)
        while True:
            # 读取帧
            ret, frame = cap.read()
            if not ret:
                break

            # 转换为灰度图像
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            # 检测二维码
            barcodes = pyzbar.decode(gray)
            for barcode in barcodes:
                # 提取条形码数据
                (x, y, w, h) = barcode.rect
                barcode_data = barcode.data.decode("utf-8")
                print(f"摄像头已识别到学生码 {barcode_data}")
                print(f"正在查询 type:{type(barcode_data)}")
                # 在图像上绘制矩形框和文本
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, barcode_data, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                if barcode_data:
                    time.sleep(1)
                    umsg=lb.Decrypt_User_Info(barcode_data)
                    print(umsg)
                    print("====student info====")
                    print(f"name:{umsg[0]}")
                    print(f"id:{umsg[1]}")
                    print(f"class:{umsg[2]}")
                    print(f"encrypt_password:{umsg[3]}")
                    print("====student info====")
                    self.info_queue.put([umsg[0], umsg[1], umsg[2], umsg[3]])
                    cap.release()  # 释放摄像头资源
                    cv2.destroyAllWindows()  # 关闭所有OpenCV窗口
                    return
    def cv_for_book(self)->list:
        global thread 
        thread = threading.Thread(target=self._cv_for_book)
        thread.start()
        thread.join()
        return self.info_queue.get()
    def _cv_for_book(self)->dict:
        global cursor1
        global thread
        print("打开摄像头扫描书籍")
                # 打开摄像头
        cap = cv2.VideoCapture(0)
        jianli_lianjie(True)
        if not cap.isOpened():
            print("Error: Could not open video source.")
            return
        while True:
            # 读取帧
            ret, frame = cap.read()
            if not ret:
                break

            # 转换为灰度图像
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            # 检测条形码
            barcodes = pyzbar.decode(gray)
            for barcode in barcodes:
                # 提取条形码数据
                (x, y, w, h) = barcode.rect
                barcode_data = barcode.data.decode("utf-8")
                print(f"摄像头已识别到图书编号 {barcode_data}")
                # 在图像上绘制矩形框和文本
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, barcode_data, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                if barcode_data:
                    time.sleep(1)
                    # 释放摄像头并关闭窗口
                    cap.release()
                    re =lb.Find_book_by_isbn(barcode_data)
                    self.info_queue.put(re)
                    cv2.destroyAllWindows()
                    return
            cv2.imshow("扫描书籍", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                cap.release()
                cv2.destroyAllWindows()
                return 0
        
    
    def delete_all_book(self):
        print(">library>delete_all_book()")
        tosql = '''
        DELETE from books
        '''
        Librarysql.execute(tosql)
        Librarysql.commit()
        print("删除所有书籍成功")

lb=Library()
#